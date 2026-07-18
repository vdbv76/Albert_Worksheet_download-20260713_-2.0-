"""
Albert Invent - Worksheet Duplicate (v3)
========================================

Faithful recreation of the Albert Worksheet in Streamlit via the official
`albert` SDK, for ANY project (MO13137 was used as the reference example).

v3 changes (driven by the "Worksheet structure" README):
  1. GLOBAL FILTER PANEL - re-implements Albert's 7 client-side UI filters
     (README §9: they are NOT persisted server-side, the API always returns
     the full set, so they must be rebuilt client-side):
        - Formula / Product ID (text)
        - Contains ingredient ("Inventory" filter - Product Design INV rows)
        - Locked / Unlocked
        - Predecessor
        - Tags
        - Data Templates
        - Created By
     One filter state -> ONE set of visible experiment columns applied to ALL
     four section tables (Product / Results / Apps / Process) AND downloads.
  2. Tags / Predecessor / Created-By come from InventoryItem entities
     (inventory.get_by_ids - batched 250/call), NOT from the grid TAG row,
     which is unreliable on large sheets (README §16.2).
  3. Data-Template membership comes from tasks.search (one call - each
     TaskSearchItem already carries its dataTemplate + inventory lists).
  4. Results are still loaded lazily per selected Property Task via
     property_data.get_all_task_properties (README §12: fastest bulk path,
     one logical call per task).
  5. interval_combination tokens (ROW4, ROW4XROW2) are resolved to human-
     readable parameter setpoints via workflows.get_by_ids ->
     Workflow.interval_combinations (README §11) - the ROW tokens are
     positional indices inside the workflow, NOT worksheet row ids.
  6. Focus view (hide rows empty across the visible columns), README §9.

v4 changes:
  1. EXACT project search - a dedicated search box filters the Project list by
     LITERAL contiguous substring ('1313' proposes 'MO13137', never '1-31-3',
     '13 13' or a lone '13'); Streamlit's fuzzy dropdown matching no longer
     decides what is proposed.
  2. 'Only my projects' now shows every project where the user is a
     collaborator in ANY role (my_role union), not only the projects they own
     (the search API's myProject flag).
  3. Interval 1 / Interval 2 in the DT panel are MULTI-selects: several
     setpoints of one axis parameter can be combined; empty = (any). Each axis
     only ever offers its own parameter's setpoints (axis identity comes from
     the workflow's intervalString), so parameters cannot be mixed.
  4. The Advanced filter's 'Results' Section entry is labelled INACTIVE until
     the Results table has been downloaded.
  5. 'Display' menu (like Albert's) - shows/hides the sheet's Blank / Lookup /
     Function / Result columns, which are now extracted from the Product grid.
  6. Reliable row tick boxes - editor state is keyed to the visible row set
     and synced via on_change, and programmatic selection changes clear the
     editor's positional deltas.
  7. The XLSX download reproduces the on-screen tables 1:1 (row filters,
     ticked rows, hidden columns, Display menu, merged vs non-merged cells per
     table, Decimals).
  8. The Results load controls (include-filtered-out, tidy view, aggregation,
     Load Data / Reload / parallel requests, load plan) moved between
     'Selection of Data Templates in Results' and the Advanced filter.

v5 changes:
  1. Project search also queries the server (projects.search text=...) so
     projects beyond the cached catalog / membership filters are found too.
  2. The results download progress bar renders directly under Load Data.
  3. Display menu rebuilt: the COMPLETE sheet-column catalog now comes from
     the raw grid's Formulas[] array (Sheet.columns only lists first-row
     cells, dropping Lookup/Function and empty Blank columns); the menu is a
     compact popover listing column HEADERS (any type), defaulting to the
     columns whose Display toggle is ON in Albert; chosen columns sit LEFT of
     the experiments, right after the key columns.
  4. Row-selection buttons regrouped into a 'Select visible rows' box
     (All / None / Apply).
  5.-7. Custom row/group reordering on every section table: an '↕ Reorder'
     popover moves whole Group/Subgroup blocks (Results: Data Template / Data
     Column; Process: Parameter group) among their siblings, and an editable
     '#' column moves a single row inside its own block - a row can never
     leave its hierarchy. Group/subgroup HEADER rows belong to their own block
     (their name is shown in their level's cell) and travel with it. The same
     ordering feeds the XLSX export.

Requirements:
    pip install streamlit albert pandas truststore openpyxl
    (Streamlit >= 1.32 recommended - st.popover; older versions fall back to
    expanders.)

Run:
    streamlit run app.py
"""

from __future__ import annotations

# --- Corporate SSL fix: must run before any HTTPS request -------------------
import truststore

truststore.inject_into_ssl()

import hashlib
import io
import os
import re
from typing import Any

import pandas as pd
import streamlit as st

from albert import Albert

st.set_page_config(page_title="Albert - Worksheet Duplicate", layout="wide")
st.title("🧪 Albert Worksheet - Live Duplicate")

SECTION_ORDER = [
    ("product_design", "Product Design"),
    ("result_design", "Results"),
    ("app_design", "Apps"),
    ("process_design", "Process Design"),
]

# Sentinel option in every filter: matches items that have NO value for that
# attribute (no tag, no predecessor, no group at that level, ...). Without it,
# selecting every real option still silently drops the blanks.
NONE_LABEL = "(None)"

ROW_TYPE_LABELS = {
    "INV": "Inventory",
    "BLK": "Blank",
    "TAS": "Task",
    "PRG": "Param. Group",
    "PRM": "Parameter",
    "TOT": "Total",
    "TAG": "Tags",
    "APP": "App link",
    "PRC": "Pricing",
    "PDC": "Predecessor",
    "BAT": "Batches",
    "LKP": "Lookup",
    "RSL": "Substance data",
    "DEF": "Default",
    "Formula": "Formula",
    "DAT": "Data Template",
    "DAC": "Data Column",
}


# ===========================================================================
# Helpers
# ===========================================================================
def _cell_text(cell: Any) -> str:
    if cell is None:
        return ""
    v = getattr(cell, "value", None)
    if v in (None, ""):
        return ""
    if isinstance(v, dict):
        for k in ("value", "name", "text"):
            if v.get(k) not in (None, ""):
                return str(v[k])
        return ""
    if isinstance(v, list):
        return ", ".join(str(x) for x in v if x not in (None, ""))
    return str(v)


def _friendly_type(raw: str) -> str:
    raw = str(raw or "")
    return ROW_TYPE_LABELS.get(raw.split(".")[-1], raw.split(".")[-1])


def _strip_inv(row_link_id: str | None) -> str:
    """Grid row ids are prefixed with 'INV'; strip to reveal the entity id
    (e.g. INVTAS123 -> TAS123)."""
    if not row_link_id:
        return ""
    return row_link_id[3:] if row_link_id.startswith("INV") else row_link_id


def _to_number(x: Any) -> tuple[Any, bool]:
    """(float, True) when x parses as a number (comma OR dot decimal), else (x, False)."""
    try:
        return float(str(x).strip().replace(",", ".")), True
    except (ValueError, AttributeError):
        return x, False


def _cmp_pass(value: Any, mode: str, a: Any, b: Any = None) -> bool:
    """Generic column filter (Product-Design fields, Results Data Column).
    Compares numerically when both sides parse as numbers, lexicographically
    otherwise. mode is one of All / > / = / < / Range (Between is a Range alias;
    Contains is still accepted for backward compatibility)."""
    if mode in (None, "", "All"):
        return True
    if a in (None, ""):
        return True
    v = str(value)
    if mode == "Contains":
        return str(a).lower() in v.lower()
    va, vnum = _to_number(v)
    aa, anum = _to_number(a)
    if mode == "=":
        return (va == aa) if (vnum and anum) else (v == str(a))
    if mode == ">":
        return (va > aa) if (vnum and anum) else (v > str(a))
    if mode == "<":
        return (va < aa) if (vnum and anum) else (v < str(a))
    if mode in ("Between", "Range"):
        if b in (None, ""):
            return True
        bb, bnum = _to_number(b)
        if vnum and anum and bnum:
            lo, hi = (aa, bb) if aa <= bb else (bb, aa)
            return lo <= va <= hi
        lo, hi = (str(a), str(b)) if str(a) <= str(b) else (str(b), str(a))
        return lo <= v <= hi
    return True


def _round_sig_decimals(tok: str, decimals: int) -> str:
    """Round ONE numeric token keeping `decimals` significant fractional digits,
    counted from the first non-zero decimal place:
        0.0012345 -> 0.0012   (decimals=2)
        1.0123    -> 1.012
        1.123     -> 1.12
    Integers and non-numeric tokens are returned unchanged."""
    t = tok.strip()
    if t == "":
        return tok
    comma = "," in t and "." not in t
    norm = t.replace(",", ".") if comma else t
    try:
        val = float(norm)
    except ValueError:
        return tok
    if "." not in norm:  # an integer -> leave it alone
        return tok
    frac = norm.split(".", 1)[1]
    lz = 0
    for ch in frac:
        if ch == "0":
            lz += 1
        else:
            break
    out = f"{val:.{lz + decimals}f}"
    return out.replace(".", ",") if comma else out


def _apply_decimals_text(text: Any, decimals: int | None) -> Any:
    """Apply `_round_sig_decimals` to every numeric token in a cell, preserving the
    ' | ' separators used to list repeated measurements."""
    if decimals is None:
        return text
    s = str(text)
    if s == "":
        return s
    if "|" not in s:
        return _round_sig_decimals(s, decimals)
    return " | ".join(_round_sig_decimals(p.strip(), decimals) for p in s.split("|"))


# ===========================================================================
# Sidebar: authentication
# ===========================================================================
with st.sidebar:
    st.header("🔐 Connect to Albert")
    base_url = st.text_input("Base URL", value="https://app.albertinvent.com")
    method = st.radio("Authentication", ["SSO login", "Client Credentials", "Static token"])
    if method == "SSO login":
        email = st.text_input("Albert account email")
        if st.button("Connect via SSO", type="primary", use_container_width=True):
            try:
                with st.spinner("Complete the login in the opened browser tab..."):
                    st.session_state["client"] = Albert.from_sso(base_url=base_url, email=email)
                st.success("Connected ✅")
            except Exception as e:  # noqa: BLE001
                st.error(f"Could not connect: {e}")
    elif method == "Client Credentials":
        cid = st.text_input("Client ID")
        sec = st.text_input("Client Secret", type="password")
        if st.button("Connect", type="primary", use_container_width=True):
            try:
                st.session_state["client"] = Albert.from_client_credentials(
                    base_url=base_url, client_id=cid, client_secret=sec
                )
                st.success("Connected ✅")
            except Exception as e:  # noqa: BLE001
                st.error(f"Could not connect: {e}")
    else:
        tok = st.text_input("Token (JWT)", type="password")
        if st.button("Connect", type="primary", use_container_width=True):
            try:
                st.session_state["client"] = Albert.from_token(base_url=base_url, token=tok)
                st.success("Connected ✅")
            except Exception as e:  # noqa: BLE001
                st.error(f"Could not connect: {e}")

client: Albert | None = st.session_state.get("client")
if client is None:
    st.info("👈 Connect to Albert from the sidebar to get started.")
    st.stop()


# ===========================================================================
# 1) Project & sheet selection - MULTI-PROJECT COMPARISON BASKET
#    The Project field IS the search box: the full project catalog is loaded
#    once (cached) and the multiselect's built-in type-ahead suggests every
#    project whose ID or description contains what the user types ('13137',
#    'MO13', '137', 'second', 'dispers', ...). Selecting a project = adding it
#    to the comparison; repeat the search to add more; the chip's ✕ removes it.
# ===========================================================================
st.header("1️⃣ Projects & sheets")

PROJECT_CATALOG_MAX = 5000

# Every ACL role a user can hold on a project (SDK AccessControlLevel).
# 'Only my projects' must show every project where the user was added as a
# collaborator REGARDLESS of the assigned role - but the search API's
# myProject flag only returns projects the user OWNS. So ownership and
# membership-in-any-role are queried separately and unioned.
PROJECT_MEMBER_ROLES = [
    "ProjectOwner",
    "ProjectEditor",
    "ProjectViewer",
    "ProjectAllTask",
    "ProjectStrictViewer",
    "ProjectPropertyTask",
]


@st.cache_data(ttl=600, show_spinner="Loading the project list from Albert...")
def load_project_catalog(_client: Albert, only_mine: bool) -> dict[str, str]:
    """label -> project id for the Project type-ahead. The label carries BOTH
    the description and the id ('<description>  [<id>]'), so typing any part
    of either filters the dropdown. `only_mine` unions the search API's
    my_project flag (projects the user OWNS) with a my_role query over every
    project role (projects where the user is a collaborator in ANY role).
    Cached for 10 min per flag value."""
    out: dict[str, str] = {}

    def _collect(**kw) -> None:
        for p in _client.projects.search(max_items=PROJECT_CATALOG_MAX, **kw):
            out.setdefault(f"{p.description or '(no description)'}  [{p.id}]", p.id)

    if only_mine:
        _collect(my_project=True)  # projects the user owns
        try:
            _collect(my_role=PROJECT_MEMBER_ROLES)  # collaborator, any role
        except Exception as e:  # noqa: BLE001
            st.warning(
                "Could not load the projects where you are a collaborator - "
                f"showing only the projects you own: {e}"
            )
    else:
        _collect()
    return out


def _proj_short(label: str, pid: str) -> str:
    """Short display code for a project, used to group Sheets / Property Tasks
    by project inside their dropdowns (e.g. 'MO13137 ▸ Sheet1'). The first
    token of the project description is normally the project code."""
    desc = label.rsplit("  [", 1)[0].strip()
    tok = desc.split()[0] if desc.split() else ""
    if 3 <= len(tok) <= 20:
        return tok
    return desc[:20] or pid


@st.cache_data(ttl=300, show_spinner=False)
def search_projects_live(_client: Albert, text: str) -> dict[str, str]:
    """Server-side full-text project search for the search box. Complements the
    cached catalog: finds projects beyond the catalog cap and projects that the
    my_project/my_role union missed (e.g. a project shared with you through a
    mechanism the membership filters don't cover). Cached 5 min per query."""
    out: dict[str, str] = {}
    try:
        for p in _client.projects.search(text=text, max_items=200):
            out[f"{p.description or '(no description)'}  [{p.id}]"] = p.id
    except Exception:  # noqa: BLE001
        pass
    return out


pc1, pc2 = st.columns([3, 1])
with pc2:
    only_mine = st.checkbox(
        "Only my projects",
        help="Suggest only projects where you are a member: projects you OWN "
        "plus every project where you were added as a collaborator, whatever "
        "the role (Editor, Viewer, All Tasks, Strict Viewer, Property Tasks).",
    )

try:
    proj_catalog = load_project_catalog(client, only_mine)
except Exception as e:  # noqa: BLE001
    st.error(f"Could not load the project list: {e}")
    st.stop()

# Every label ever offered, so a project selected earlier stays resolvable
# (and selected) even when 'Only my projects' later hides it from the catalog.
_label_to_id: dict[str, str] = st.session_state.setdefault("proj_label_to_id", {})
_label_to_id.update(proj_catalog)

# KEYED widget + shadow copy. A stable `key` is what makes CONSECUTIVE
# selection changes stick: the previous pattern (auto-keyed widget fed through
# `default=`) changes the widget's identity on every rerun that follows a
# selection change, so the very next change was silently dropped. The shadow
# (plain session key) restores the basket if some st.rerun() ever aborts the
# script before this widget runs (Streamlit then garbage-collects keyed
# widget state).
_PROJ_KEY = "proj_basket_widget"
if _PROJ_KEY not in st.session_state:
    st.session_state[_PROJ_KEY] = list(st.session_state.get("proj_basket", []))
st.session_state[_PROJ_KEY] = [
    l for l in st.session_state[_PROJ_KEY] if l in _label_to_id
]
_proj_options = [
    l for l in st.session_state[_PROJ_KEY] if l not in proj_catalog
] + list(proj_catalog)

with pc1:
    # EXACT (contiguous) search box. Streamlit's own dropdown type-ahead does
    # FUZZY matching - typing '1313' also proposes '1-31-3', '13 13' or a lone
    # '13' because it matches the characters in order but not necessarily
    # adjacent. This box pre-filters the options with a literal substring
    # test, so only projects whose ID or description contains '1313'
    # contiguously are proposed. Already-selected projects always stay in the
    # option list so their chips never break.
    proj_query = st.text_input(
        "Search project (exact match)",
        key="proj_search",
        placeholder="e.g. 1313 → only projects containing '1313' together",
        help="Literal, contiguous match on the project ID or description: "
        "'1313' proposes 'MO13137' but never '1-31-3', '13 13' or a lone '13'.",
    )
    _q = proj_query.strip().lower()
    if _q:
        _selected_now = set(st.session_state[_PROJ_KEY])
        # ALSO search server-side: catches projects beyond the catalog cap and
        # projects the membership filters missed (e.g. EXP13976 - editor role
        # but not returned by myProject). The literal-substring test is still
        # applied to everything that comes back.
        _live = search_projects_live(client, proj_query.strip())
        _label_to_id.update(_live)
        _merged = list(dict.fromkeys(list(_proj_options) + list(_live)))
        _proj_options = [
            l for l in _merged if _q in l.lower() or l in _selected_now
        ]
    sel_proj_labels = st.multiselect(
        "Project",
        _proj_options,
        key=_PROJ_KEY,
        placeholder="Pick a project (use the search box above to narrow the list)...",
        help="Use the search box above for an exact (contiguous) match, then "
        "pick the project here. Repeat the search to add the next project: "
        "every selected project joins the comparison. Remove one with the ✕ "
        "on its chip.",
    )
st.session_state["proj_basket"] = sel_proj_labels

if len(proj_catalog) >= PROJECT_CATALOG_MAX:
    st.caption(
        f"⚠️ Project list capped at the {PROJECT_CATALOG_MAX} most recent projects - "
        "a project missing from the suggestions may be beyond the cap."
    )

if not sel_proj_labels:
    st.info("Select at least one project to load its worksheet.")
    st.stop()

# [(label, project_id, short_code), ...] in selection order
selected_projects = [
    (l, _label_to_id[l], _proj_short(l, _label_to_id[l])) for l in sel_proj_labels
]
MULTI_PROJECT = len(selected_projects) > 1
PROJECT_IDS = [pid for _, pid, _ in selected_projects]
COMPARE_TAG = "_".join(s for _, _, s in selected_projects)[:80]
TITLE_PROJECTS = "  +  ".join(l for l, _, _ in selected_projects)


@st.cache_resource(show_spinner="Loading worksheet from Albert...")
def get_worksheet(_client: Albert, pid: str):
    return _client.worksheets.get_by_project_id(project_id=pid)


# --- one grouped Sheet dropdown across every selected project ---------------
# Streamlit multiselects have no native option groups, so grouping is done by
# ordering (all of project 1's sheets, then project 2's, ...) and by prefixing
# each sheet with its project code: 'MO13137 ▸ Sheet1'.
sheet_entries: dict[str, tuple[str, str, str, Any]] = {}  # label -> (pid, short, sheet_name, sheet_obj)
for _plabel, _pid, _pshort in selected_projects:
    try:
        _ws = get_worksheet(client, _pid)
    except Exception as e:  # noqa: BLE001
        st.error(f"Could not load the worksheet of {_plabel}: {e}")
        continue
    _shs = getattr(_ws, "sheets", None) or []
    if not _shs:
        st.warning(f"{_pshort}: this project's worksheet has no sheets.")
        continue
    for _s in _shs:
        _nm = getattr(_s, "name", "") or "(unnamed)"
        _lbl = f"{_pshort} ▸ {_nm}"
        _n = 2
        while _lbl in sheet_entries:  # two projects may share code + sheet name
            _lbl = f"{_pshort} ▸ {_nm} ({_n})"
            _n += 1
        sheet_entries[_lbl] = (_pid, _pshort, _nm, _s)

if not sheet_entries:
    st.stop()

# Default: the FIRST sheet of a newly added project is auto-selected once, so
# a fresh comparison shows data immediately - but a sheet the user deselects
# afterwards stays deselected. Keyed widget + shadow, same reasoning as the
# Project picker above (consecutive changes + rerun-abort survival).
_SHEET_KEY = "sheets_basket_widget"
if _SHEET_KEY not in st.session_state:
    st.session_state[_SHEET_KEY] = list(st.session_state.get("sheets_basket", []))
_seeded: set[str] = st.session_state.setdefault("sheets_seeded", set())
_prev_sheets = [l for l in st.session_state[_SHEET_KEY] if l in sheet_entries]
for _plabel, _pid, _pshort in selected_projects:
    if _pid in _seeded:
        continue
    _first = next((lbl for lbl, v in sheet_entries.items() if v[0] == _pid), None)
    if _first:
        _prev_sheets.append(_first)
        _seeded.add(_pid)
st.session_state[_SHEET_KEY] = list(dict.fromkeys(_prev_sheets))

sel_sheet_labels = st.multiselect(
    "Sheet",
    list(sheet_entries.keys()),
    key=_SHEET_KEY,
    help="Sheets are grouped by project (PROJECT ▸ Sheet). Select any number - "
    "the selected sheets are merged into ONE comparison worksheet below: rows "
    "with the same Group / name align on one row, and every project's "
    "experiment columns appear side by side.",
)
st.session_state["sheets_basket"] = sel_sheet_labels

if not sel_sheet_labels:
    st.info("Select at least one sheet to compare.")
    st.stop()

TITLE_SHEETS = "  |  ".join(sel_sheet_labels)


# ===========================================================================
# 2) Extract sheet structure (grid values + column metadata incl. locked)
# ===========================================================================
@st.cache_data(show_spinner="Reading sheet grid...")
def extract_sheet(_sheet, sheet_key: str) -> dict:
    def _row_hierarchy(design) -> dict:
        """rowId -> ordered ancestor names [Group, Subgroup 1, ..., Subgroup n].

        SOURCE OF TRUTH: GET /api/v3/worksheet/design/{id}/rows/sequence.
        Row grouping in Albert is an explicit server-side parent->child
        relationship (Design.group_rows() PUTs an explicit ChildRows list), so
        the tree is stored, not inferred. The SDK's own get_groups() reads only
        ONE level of that response and throws away nested subgroups - so we
        call the endpoint directly and walk it recursively to arbitrary depth.

        The grid/sheet_inspect endpoint is a FLAT array with no parent, child,
        depth or indent field. Depth CANNOT be recovered from it: a BLK row
        followed by another BLK row is ambiguous between "child" and "sibling",
        and that ambiguity is not resolvable by any rule. We therefore do NOT
        guess. If the sequence endpoint is unavailable, hierarchy is reported
        as UNAVAILABLE rather than fabricated.

        Returns {"paths": {rid: [ancestor_rid, ...]}, "node_names": {rid: name},
                 "source": str, "raw": <json|None>, "error": str|None, "keys": {...}}

        NOTE: the sequence tree identifies nodes by rowId and does not reliably
        carry a display name. Ancestor NAMES are therefore resolved afterwards
        from the grid's rowId -> name map (BLK group headers appear in the grid
        with their labels, e.g. ROW338 -> "Raw Materials"). Any name the tree
        does supply is kept in node_names and wins over the grid.
        """
        out = {
            "paths": {},
            "node_names": {},
            "source": "unavailable",
            "raw": None,
            "error": None,
            "keys": {},
        }

        seq = None
        try:
            resp = design.session.get(f"/api/v3/worksheet/design/{design.id}/rows/sequence")
            seq = resp.json()
            out["raw"] = seq
        except Exception as e:  # noqa: BLE001
            out["error"] = f"{type(e).__name__}: {e}"

        ID_KEYS = ("rowId", "row_id", "id")
        NAME_KEYS = ("name", "lableName", "labelName", "rowName", "label")
        KID_KEYS = ("children", "childRows", "ChildRows", "Children", "Rows", "rows")

        def first(node: dict, keys) -> Any:
            for k in keys:
                v = node.get(k)
                if v not in (None, "", []):
                    out["keys"][k] = out["keys"].get(k, 0) + 1
                    return v
            return None

        def kids(node: dict) -> list:
            for k in KID_KEYS:
                v = node.get(k)
                if isinstance(v, list) and v:
                    out["keys"][k] = out["keys"].get(k, 0) + 1
                    return v
            return []

        paths: dict[str, list[str]] = {}
        nested = False

        def walk(nodes, ancestor_ids: list[str]):
            nonlocal nested
            for n in nodes:
                if not isinstance(n, dict):
                    continue
                rid = first(n, ID_KEYS)
                if not rid:
                    continue
                rid = str(rid)
                nm = first(n, NAME_KEYS)
                if nm:
                    out["node_names"][rid] = str(nm)
                paths[rid] = ancestor_ids[:]
                ch = kids(n)
                if ch:
                    nested = True
                    walk(ch, ancestor_ids + [rid])

        if isinstance(seq, list) and seq:
            walk(seq, [])
        elif isinstance(seq, dict):
            for k in ("Items", "items", "Rows", "rows", "data", "Data"):
                if isinstance(seq.get(k), list):
                    walk(seq[k], [])
                    break

        if paths and nested:
            out["paths"] = paths
            out["source"] = "rows/sequence endpoint (explicit parent->child tree)"
            return out

        # The endpoint returned nothing usable. Last resort: the SDK's
        # one-level get_groups() (same endpoint, shallow parse) - gives a
        # single Group level only, never subgroups. Still real data, not a guess.
        try:
            groups = design.get_groups()
        except Exception:  # noqa: BLE001
            groups = []
        if groups:
            for g in groups:
                if g.name:
                    out["node_names"][g.row_id] = g.name
                for ch in g.child_row_ids:
                    paths[ch] = [g.row_id]
            out["paths"] = paths
            out["source"] = "get_groups() - ONE LEVEL ONLY (no subgroups available)"
            return out

        out["source"] = "unavailable"
        return out

    inv_to_form_name = {
        f.id: (f.name or "")
        for f in (getattr(_sheet, "formulations", None) or [])
        if getattr(f, "id", None)
    }

    columns = []
    for c in _sheet.columns:
        name = getattr(c, "name", None) or ""
        inv_id = getattr(c, "inventory_id", None)
        columns.append(
            {
                "column_id": getattr(c, "column_id", None),
                "name": name,
                "type": str(getattr(c, "type", "") or ""),
                "inventory_id": inv_id,
                "col_key": inv_id,
                "hidden": bool(getattr(c, "hidden", False)),
                "locked": bool(getattr(c, "locked", False)),
                "pinned": getattr(c, "pinned", None),
                "formulation_name": inv_to_form_name.get(inv_id, ""),
                # The sheet's built-in label column duplicates the row names -
                # exclude it from the data columns.
                "is_label_col": name.strip().lower() == "name",
            }
        )

    # ---- COMPLETE special-column catalog (Blank / Lookup / Function / ...) ----
    # The SDK's Sheet.columns is built from the FIRST ROW's cells only, so it
    # misses every Lookup/Function column and every Blank column with no value
    # on that row. The authoritative column list is the raw grid response's
    # Formulas[] array (one entry per column, with name + hidden - hidden is
    # exactly Albert's Display-menu toggle state). One extra GET per sheet,
    # cached with the rest of this extraction. Special columns are keyed
    # 'COL::<sheet>::<colId>' (no inventory_id), unique across a comparison.
    _pd_design = getattr(_sheet, "product_design", None)
    special_catalog: list[dict] = []
    if _pd_design is not None:
        raw_grid = None
        try:
            resp = _pd_design.session.get(
                f"/api/v3/worksheet/{_pd_design.id}/{_pd_design.design_type.value}/grid"
            )
            raw_grid = resp.json()
        except Exception:  # noqa: BLE001
            raw_grid = None
        if isinstance(raw_grid, dict):
            # cell type / header name per column, scanned across EVERY row
            type_of: dict[str, str] = {}
            name_of: dict[str, str] = {}
            for item in raw_grid.get("Items") or []:
                for v in item.get("Values") or []:
                    cid = v.get("colId")
                    if not cid:
                        continue
                    t = str(v.get("type") or "")
                    if t and cid not in type_of:
                        type_of[cid] = t
                    nm = v.get("name")
                    if nm and cid not in name_of:
                        name_of[cid] = str(nm)
            inv_col_ids = {c["column_id"] for c in columns if c["inventory_id"]}
            for f in raw_grid.get("Formulas") or []:
                cid = f.get("colId") if isinstance(f, dict) else None
                if not cid or cid in inv_col_ids:
                    continue
                tcode = type_of.get(cid, "").split(".")[-1]
                if tcode in ("INV", "Formula", "TAG"):
                    continue  # formulation / tag columns are handled above
                nm = str(f.get("name") or name_of.get(cid) or "")
                if nm.strip().lower() == "name":
                    continue  # the built-in label column
                special_catalog.append(
                    {
                        "column_id": cid,
                        "name": nm or cid,
                        "type": tcode or "BLK",
                        "inventory_id": None,
                        "col_key": f"COL::{sheet_key}::{cid}",
                        # Formulas[].hidden IS the Display-menu toggle in Albert
                        "hidden": bool(f.get("hidden") or False),
                        "locked": False,
                        "pinned": None,
                        "formulation_name": "",
                        "is_label_col": False,
                    }
                )
    # drop Sheet.columns duplicates of the special columns, then append the
    # complete catalog (the catalog entry carries the authoritative hidden flag)
    _special_cids = {c["column_id"] for c in special_catalog}
    columns = [
        c
        for c in columns
        if c["inventory_id"] or c["is_label_col"] or c["column_id"] not in _special_cids
    ]
    columns.extend(special_catalog)

    # column_id -> COL key, for capturing the special columns' cell values
    # (valid on the PRODUCT grid only - other designs' column_ids differ).
    special_key_of = {c["column_id"]: c["col_key"] for c in special_catalog}

    sections = []
    for attr, label in SECTION_ORDER:
        design = getattr(_sheet, attr, None)
        if design is None:
            continue
        try:
            grid = design.grid
        except Exception:  # noqa: BLE001
            continue
        if grid is None or grid.empty:
            continue

        # FIX: each Design parses its OWN columns from its OWN grid response
        # (Process Design even uses a different endpoint), so column_ids are NOT
        # comparable across designs - Sheet.columns is product_design.columns only.
        # Key every cell by inventory_id, which IS stable across designs. This is
        # what makes one filter state apply identically to all four sections.
        try:
            design_col_inv = {
                getattr(c, "column_id", None): getattr(c, "inventory_id", None)
                for c in design.columns
            }
        except Exception:  # noqa: BLE001
            design_col_inv = {}

        # Full nested hierarchy (Group / Subgroup 1 / ... / Subgroup n)
        hier = _row_hierarchy(design)
        row_paths = hier["paths"]

        rows = []
        for _, row_series in grid.iterrows():
            first_cell = next((c for _, c in row_series.items() if c is not None), None)
            if first_cell is None:
                continue
            rid = getattr(first_cell, "row_id", None)
            label_name = (
                getattr(first_cell, "row_label_name", None)
                or getattr(first_cell, "name", None)
                or ""
            )
            rtype_raw = str(getattr(first_cell, "row_type", "") or "")
            link_id = _strip_inv(getattr(first_cell, "inventory_id", None))
            values: dict[str, str] = {}  # inventory_id -> cell text
            for _, cell in row_series.items():
                if cell is None or (isinstance(cell, float) and pd.isna(cell)):
                    continue
                cid = getattr(cell, "column_id", None)
                # NB: cell.inventory_id is the ROW's inventory item (the raw
                # material), never the column's - only the design column map is valid.
                inv = design_col_inv.get(cid)
                if inv:
                    values[inv] = _cell_text(cell)
                elif attr == "product_design" and cid in special_key_of:
                    # Blank / Lookup / Function / Result column cell (Display menu)
                    values[special_key_of[cid]] = _cell_text(cell)
            path_ids = list(row_paths.get(rid, []))  # [ancestor rowIds, outer->inner]
            rows.append(
                {
                    "row_id": rid,
                    "name": label_name,
                    "type_raw": rtype_raw,
                    "type": _friendly_type(rtype_raw),
                    "path_ids": path_ids,
                    "path": [],  # filled in below, once every row name is known
                    "depth": len(path_ids) + 1,
                    "link_id": link_id,  # e.g. TAS123 for Property Block rows
                    "values": values,
                }
            )

        # --- resolve ancestor rowIds -> display names -------------------------
        # The sequence tree is keyed by rowId and carries no reliable label, so
        # the group/subgroup names come from the grid itself: a group header IS
        # a BLK row, and that row's name is its label (ROW338 -> "Raw Materials").
        name_by_rid = {r["row_id"]: r["name"] for r in rows if r["row_id"]}
        node_names = hier["node_names"]  # names the tree did supply, if any
        unresolved: set[str] = set()
        for r in rows:
            names = []
            for aid in r["path_ids"]:
                nm = node_names.get(aid) or name_by_rid.get(aid) or ""
                if not nm:
                    unresolved.add(aid)
                    nm = str(aid)  # last resort: show the id rather than blank
                names.append(nm)
            r["path"] = names

        if rows:
            sections.append(
                {
                    "attr": attr,
                    "label": label,
                    "rows": rows,
                    "hierarchy_source": hier["source"],
                    "hierarchy_error": hier["error"],
                    "hierarchy_keys": hier["keys"],
                    "hierarchy_raw": hier["raw"],
                    "hierarchy_unresolved": sorted(unresolved),
                    "max_depth": max((len(r["path"]) for r in rows), default=0),
                }
            )

    return {"columns": columns, "sections": sections}


def merge_extracted(entries: list[dict]) -> dict:
    """Fold every selected (project, sheet) into ONE virtual worksheet.

    COLUMNS are concatenated in selection order (project 1's sheets first),
    deduped by inventory_id (the same formulation on two selected sheets keeps
    one column), and each carries an `origin` tag ('MO13137 / Sheet1') that
    rides along in header tooltips and the description row.

    ROWS are aligned across sheets by (ancestor path, name, row type): the
    same ingredient / parameter row in two projects lands on ONE row whose
    `values` dict (keyed by inventory_id, globally unique) holds every
    project's cells side by side. Rows unique to a later sheet are appended
    after the earlier sheets' rows, keeping each sheet's own order intact.
    row_ids are prefixed with the project id - grid row ids like ROW4 repeat
    across projects and would otherwise collide in the selection state.
    """
    multi = len(entries) > 1
    columns: list[dict] = []
    seen_inv: set[str] = set()
    for e in entries:
        origin = f"{e['short']} / {e['sheet']}" if multi else ""
        for c in e["data"]["columns"]:
            inv = c["inventory_id"]
            if inv:
                if inv in seen_inv:
                    continue
                seen_inv.add(inv)
            cc = dict(c)
            cc["origin"] = origin
            columns.append(cc)

    merged: dict[str, dict] = {}
    for e in entries:
        for s in e["data"]["sections"]:
            tgt = merged.get(s["attr"])
            if tgt is None:
                tgt = {
                    "attr": s["attr"],
                    "label": s["label"],
                    "rows": [],
                    "hierarchy_source": [],
                    "hierarchy_error": None,
                    "hierarchy_keys": {},
                    "hierarchy_raw": {},
                    "hierarchy_unresolved": set(),
                    "max_depth": 0,
                    "_row_index": {},
                }
                merged[s["attr"]] = tgt
            tgt["hierarchy_source"].append(f"{e['label']}: {s['hierarchy_source']}")
            tgt["hierarchy_error"] = tgt["hierarchy_error"] or s["hierarchy_error"]
            for k, n in (s["hierarchy_keys"] or {}).items():
                tgt["hierarchy_keys"][k] = tgt["hierarchy_keys"].get(k, 0) + n
            if s["hierarchy_raw"] is not None:
                tgt["hierarchy_raw"][e["label"]] = s["hierarchy_raw"]
            tgt["hierarchy_unresolved"].update(s.get("hierarchy_unresolved") or [])
            tgt["max_depth"] = max(tgt["max_depth"], s["max_depth"])
            for r in s["rows"]:
                key = (tuple(r["path"]), r["name"], r["type_raw"].split(".")[-1])
                hit = tgt["_row_index"].get(key)
                if hit is None:
                    rr = dict(r)
                    rr["row_id"] = f"{e['pid']}::{r['row_id']}"
                    rr["values"] = dict(r["values"])
                    rr["origins"] = [e["label"]]
                    tgt["_row_index"][key] = rr
                    tgt["rows"].append(rr)
                else:
                    hit["values"].update(r["values"])
                    hit["origins"].append(e["label"])

    sections = []
    for attr, _lab in SECTION_ORDER:
        sec = merged.get(attr)
        if not sec:
            continue
        sec.pop("_row_index")
        sec["hierarchy_source"] = " | ".join(sec["hierarchy_source"])
        sec["hierarchy_raw"] = sec["hierarchy_raw"] or None
        sec["hierarchy_unresolved"] = sorted(sec["hierarchy_unresolved"])
        sections.append(sec)
    return {"columns": columns, "sections": sections}


sheet_data_entries: list[dict] = []
for _lbl in sel_sheet_labels:
    _pid, _pshort, _snm, _sobj = sheet_entries[_lbl]
    try:
        _d = extract_sheet(_sobj, f"{_pid}::{_snm}")
    except Exception as e:  # noqa: BLE001
        st.error(f"Could not read {_lbl}: {e}")
        continue
    sheet_data_entries.append(
        {"pid": _pid, "short": _pshort, "sheet": _snm, "label": _lbl, "data": _d}
    )

if not sheet_data_entries:
    st.warning("None of the selected sheets could be read.")
    st.stop()

data = merge_extracted(sheet_data_entries)
columns, sections = data["columns"], data["sections"]

if len(sheet_data_entries) > 1:
    st.caption(
        "🔀 **Comparing "
        + " · ".join(f"`{e['label']}`" for e in sheet_data_entries)
        + "** - merged into one worksheet: same-named rows are aligned on one row, "
        "each project keeps its own experiment columns."
    )

if not sections:
    st.warning("No grid data found in the selected sheet(s).")
    st.stop()

section_by_attr = {s["attr"]: s for s in sections}


# ===========================================================================
# 3) Enrichment - the data behind the filters
#    (README §9/§14/§16: tags, predecessor and creator come from Inventory
#    entities, NOT the grid; Data-Template membership comes from task search)
# ===========================================================================
def _tag_name(t: Any) -> str:
    """A Tag's NAME lives in `Tag.tag` (alias 'name'/'tagName'), NOT `Tag.name`.
    Reading `.name` silently yields None for every tag - which is why the Tags
    filter came up empty. Falls back through the other shapes defensively."""
    if isinstance(t, str):
        return t
    for attr in ("tag", "name", "tag_name"):
        v = getattr(t, attr, None)
        if v:
            return str(v)
    if isinstance(t, dict):
        for k in ("tag", "name", "tagName"):
            if t.get(k):
                return str(t[k])
    return ""


def _tag_id(t: Any) -> str:
    for attr in ("id", "albert_id", "tag_id"):
        v = getattr(t, attr, None)
        if v:
            return str(v)
    if isinstance(t, dict):
        return str(t.get("id") or t.get("albertId") or "")
    return ""


@st.cache_data(show_spinner="Loading formulation metadata (tags, creators)...")
def load_inventory_meta(_client: Albert, inv_ids: tuple[str, ...]) -> dict[str, dict]:
    """inventory_id -> {name, alias, description, tags, created_by}.

    TAGS: `get_by_ids` returns tag links that may carry only the TAG id, so any
    id whose name is missing is resolved in one batched `tags.get_by_ids` call.
    PREDECESSOR is NOT on the InventoryItem at all (no top-level field, not in
    Metadata, no facet) - it comes from the worksheet's Apps PDC row instead;
    see `predecessor_by_inv()`.
    """
    out: dict[str, dict] = {}
    ids = [i for i in inv_ids if i]
    if not ids:
        return out
    try:
        items = _client.inventory.get_by_ids(ids=list(ids))
    except Exception as e:  # noqa: BLE001
        st.warning(f"Could not load inventory metadata (filters degraded): {e}")
        return out

    unresolved_tag_ids: set[str] = set()
    for it in items:
        tag_names, tag_ids = [], []
        for t in getattr(it, "tags", None) or []:
            nm, tid = _tag_name(t), _tag_id(t)
            if nm:
                tag_names.append(nm)
            elif tid:
                tag_ids.append(tid)
                unresolved_tag_ids.add(tid)
        created = getattr(it, "created", None)
        out[it.id] = {
            "name": getattr(it, "name", "") or "",
            "alias": getattr(it, "alias", "") or "",
            "description": getattr(it, "description", "") or "",
            "tags": tag_names,
            "_tag_ids": tag_ids,  # names still to resolve
            "created_by": getattr(created, "by_name", None) or getattr(created, "by", "") or "",
            "created_at": str(getattr(created, "at", "") or ""),
        }

    # Resolve any id-only tags -> names (one batched call)
    if unresolved_tag_ids:
        id_to_name: dict[str, str] = {}
        try:
            for t in _client.tags.get_by_ids(ids=sorted(unresolved_tag_ids)):
                id_to_name[_tag_id(t)] = _tag_name(t)
        except Exception:  # noqa: BLE001
            pass
        for m in out.values():
            for tid in m["_tag_ids"]:
                nm = id_to_name.get(tid)
                m["tags"].append(nm if nm else tid)

    for m in out.values():
        m["tags"] = sorted(set(t for t in m["tags"] if t))
        m.pop("_tag_ids", None)
    return out


@st.cache_data(show_spinner="Loading filter facets...")
def load_facets(_client: Albert, pid: str) -> dict[str, list[tuple[str, int]]]:
    """Albert's own filter-dropdown source: inventory facets, project-scoped.
    parameter -> [(value, count), ...]. Used to seed the Tags / Created By
    dropdowns so they show every value that exists, with counts."""
    out: dict[str, list[tuple[str, int]]] = {}
    try:
        for f in _client.inventory.get_all_facets(project_id=pid):
            out[f.parameter] = [(v.name, v.count) for v in (f.value or [])]
    except Exception:  # noqa: BLE001
        pass
    return out


# ===========================================================================
# DataTemplate-first Results: DT index built from tasks.get_all inline Blocks.
#
# STEP-0 SDK VERIFICATION (albert v1.34.0, via inspect/model_fields - findings):
#   1. DT definition accessor is `client.data_templates.get_by_id(id=<DAT id>)`.
#      The returned DataTemplate exposes `data_column_values` (alias
#      "DataColumns"); each DataColumnValue carries `data_column_id` (alias
#      "id"), `name`, `hidden: bool`, `sequence`, and `unit` (alias "Unit").
#      `Unit.symbol` exists on the full Unit model, BUT the field is typed
#      `Unit | EntityLink` - an EntityLink has no symbol - so the accessor
#      below reads symbol defensively (symbol -> name -> "").
#   2. `PropertyTask.blocks` (alias "Blocks") is `list[Block] | None`.
#      `Block.workflow` (alias "Workflow") is ALWAYS a list.
#      `Block.data_template` (alias "Datatemplate") is NOT always a list:
#      typed `list[BlockDataTemplateInfo] | DataTemplateAndTargets |
#      list[DataTemplate | EntityLink]` -> the `_first()` helper handles both.
#   3. Nothing in the model forbids two blocks with the SAME dataTemplateId in
#      one task, so the index maps DT -> list[(task, block, workflow)] - never 1:1.
#   Also confirmed: `tasks.get_all(project_id=..., category=...)` takes a
#   SINGULAR project_id (loop per project) and its `data_template=` filter is
#   by NAME (not used - filtering is done client-side on DAT ids).
# ===========================================================================
def _first(x: Any) -> Any:
    """Block.data_template / Block.workflow may be a list or a bare object."""
    if isinstance(x, (list, tuple)):
        return x[0] if x else None
    return x


@st.cache_data(show_spinner="Indexing Property Tasks & Data Templates...")
def load_property_task_catalog(
    _client: Albert, project_ids: tuple[str, ...], cache_bust: int
) -> dict:
    """ONE `tasks.get_all(project_id=p, category='Property')` call PER project
    (project_id is singular). The full task objects carry an INLINE Blocks[]
    array (block id + Datatemplate + Workflow) and the task-level Inventories[]
    - no per-task get_by_id needed and ZERO property data downloaded.

    Returns {"tasks": [...], "dt_index": {...}} where
        dt_index[dt_id] = {"name": str,
                           "occurrences": [{"task_id", "block_id",
                                            "workflow_id", "project_id",
                                            "task_name"}, ...]}
    `cache_bust` only busts the st.cache_data entry (🔄 Reload)."""
    tasks_out: list[dict] = []
    dt_index: dict[str, dict] = {}
    for pid in project_ids:
        try:
            found = list(
                _client.tasks.get_all(project_id=pid, category="Property", max_items=1000)
            )
        except Exception as e:  # noqa: BLE001
            st.warning(f"Task listing failed for {pid}: {e}")
            continue
        for t in found:
            tid = getattr(t, "id", None)
            tname = getattr(t, "name", "") or ""
            if not tid:
                continue
            blocks_out: list[dict] = []
            for b in getattr(t, "blocks", None) or []:
                bid = getattr(b, "id", None)
                dt = _first(getattr(b, "data_template", None))
                wf = _first(getattr(b, "workflow", None))
                dt_id = getattr(dt, "id", None)
                dt_name = getattr(dt, "name", None) or getattr(dt, "full_name", None) or ""
                wf_id = getattr(wf, "id", None) or ""
                if not (bid and dt_id):
                    continue
                blocks_out.append(
                    {"block_id": bid, "dt_id": dt_id, "dt_name": dt_name, "workflow_id": wf_id}
                )
                ent = dt_index.setdefault(dt_id, {"name": dt_name, "occurrences": []})
                if dt_name and not ent["name"]:
                    ent["name"] = dt_name
                ent["occurrences"].append(
                    {
                        "task_id": tid,
                        "block_id": bid,
                        "workflow_id": wf_id,
                        "project_id": pid,
                        "task_name": tname,
                    }
                )
            inv_ids = [
                getattr(ii, "inventory_id", None)
                for ii in (getattr(t, "inventory_information", None) or [])
            ]
            tasks_out.append(
                {
                    "id": tid,
                    "name": tname,
                    "state": str(getattr(t, "state", "") or ""),
                    "project_id": pid,
                    "data_templates": [b["dt_name"] for b in blocks_out if b["dt_name"]],
                    "data_template_ids": [b["dt_id"] for b in blocks_out],
                    "inventory_ids": [i for i in inv_ids if i],
                    "blocks": blocks_out,
                }
            )
    return {"tasks": tasks_out, "dt_index": dt_index}


@st.cache_data(show_spinner=False)
def load_dt_definition(_client: Albert, dt_id: str, cache_bust: int) -> list[dict]:
    """DataColumns of ONE Data Template, from `data_templates.get_by_id`.
    Property data returns only `Unit.id` - the display SYMBOL exists only here,
    so it is captured (defensively: the model allows a bare EntityLink)."""
    out: list[dict] = []
    try:
        dt = _client.data_templates.get_by_id(id=dt_id)
    except Exception as e:  # noqa: BLE001
        st.warning(f"Could not load Data Template {dt_id}: {e}")
        return out
    for dcv in getattr(dt, "data_column_values", None) or []:
        u = getattr(dcv, "unit", None)
        out.append(
            {
                "id": getattr(dcv, "data_column_id", None) or "",
                "name": getattr(dcv, "name", "") or "",
                "hidden": bool(getattr(dcv, "hidden", False)),
                "sequence": str(getattr(dcv, "sequence", "") or ""),
                "unit_id": getattr(u, "id", None) or "",
                "unit_symbol": getattr(u, "symbol", None) or getattr(u, "name", None) or "",
            }
        )
    return [d for d in out if d["id"]]


@st.cache_data(show_spinner="Resolving data template names...")
def load_data_template_names(_client: Albert, dt_ids: tuple[str, ...]) -> dict[str, str]:
    """DataTemplate id -> canonical short name. The name string carried on the
    task-search response is not necessarily the one the Worksheet dropdown shows
    (DataTemplate has name / fullName / originalName, e.g. DAT235 is 'Cobb Value'
    but fullName is 'DIN EN 20535: Cobb Value'). Resolve from the entity itself."""
    out: dict[str, str] = {}
    ids = [i for i in dt_ids if i]
    if not ids:
        return out
    try:
        for dt in _client.data_templates.get_by_ids(ids=list(ids)):
            full = getattr(dt, "full_name", None)
            nm = getattr(dt, "name", "") or ""
            out[dt.id] = f"{nm} ({full})" if full and full != nm else nm
    except Exception:  # noqa: BLE001
        pass
    return out


# --- interval resolver (moved above the Filters section: the DT selectors
# need it at render time; body unchanged) --------------------------------
ROW_TOKEN_RE = re.compile(r"ROW\d+")


def _workflow_raw(_client: Albert, wf_id: str) -> dict | None:
    """Raw workflow JSON, bypassing the SDK's typed model.

    Some workflows in this tenant carry placeholder intervals that have a Unit
    but no `value`. The SDK's Interval model marks `value` as required, so
    `workflows.get_by_id()` raises a pydantic ValidationError on an otherwise
    well-formed response - and with `get_by_ids()` one bad workflow kills the whole
    batch. So we go to the wire ourselves and read the dict."""
    try:
        resp = _client.session.get(f"/api/v3/workflows/{wf_id}")
        data = resp.json()
    except Exception:  # noqa: BLE001
        return None
    if isinstance(data, dict) and "Items" in data:
        items = data.get("Items") or []
        return items[0] if items else None
    return data if isinstance(data, dict) else None


def _workflow_interval_map(_client: Albert, wf_id: str) -> dict[str, list[str]]:
    """token -> ordered axis setpoints, e.g.
       "ROW3XROW22" -> ["Time: 0 day", "Speed: 20 RPM"]

    SOURCE OF TRUTH: `IntervalCombinations[]`, where each entry has
        interval        "ROW3XROW22"                  <- the token on property data
        intervalParams  "INT1XINT1"                   <- per-parameter 1-based index
        intervalString  "Time: 0 day,Speed: 20 RPM"   <- RESOLVED, ORDERED setpoints

    The three fields are positionally aligned, so the Nth ROW token belongs to the
    Nth comma-segment of intervalString. Axis identity is READ, never assumed.

    Two things I previously got wrong, both disproved by the live payload:
      * `Parameters[].Intervals[].rowId` is NOT a reliable source - on WFL446095 the
        interval objects carry no rowId at all, so that map came out empty and every
        token stayed raw.
      * A ROW token is NOT a positional index and NOT the parameter's own
        `prgPrmRowId`: Time's parameter row is ROW1, yet its four interval values
        carry the tokens ROW3-ROW6. Tokens are workflow-scoped ids per interval VALUE.
    """
    out: dict[str, list[str]] = {}

    combos: list = []
    try:  # typed path first
        wf = _client.workflows.get_by_id(id=wf_id)
        for ic in getattr(wf, "interval_combinations", None) or []:
            combos.append(
                {
                    "interval": getattr(ic, "interval_id", None),
                    "intervalString": getattr(ic, "interval_string", None),
                    "intervalDetails": getattr(ic, "interval_details", None),
                }
            )
    except Exception:  # noqa: BLE001  (ValidationError on placeholder intervals)
        raw = _workflow_raw(_client, wf_id)
        if raw:
            combos = raw.get("IntervalCombinations") or []

    for c in combos:
        token = c.get("interval") if isinstance(c, dict) else None
        if not token:
            continue
        token = str(token)
        # X-SAFE: never str.split("X") - a unit or parameter name may contain an X.
        # ROW tokens are always ROW<int>, so a regex findall is safe.
        row_tokens = ROW_TOKEN_RE.findall(token)

        istr = c.get("intervalString") if isinstance(c, dict) else None
        axes: list[str] = []
        if istr:
            # split on "," between axes, then the FIRST ":" between name and value
            for seg in str(istr).split(","):
                name, _, val = seg.partition(":")
                name, val = name.strip(), val.strip()
                axes.append(f"{name}: {val}" if name and val else (val or name))
        else:
            details = c.get("intervalDetails") or []
            for d in details:
                nm = d.get("name") if isinstance(d, dict) else getattr(d, "name", "")
                vl = d.get("value") if isinstance(d, dict) else getattr(d, "value", "")
                axes.append(f"{nm}: {vl}".strip(": "))

        # N-AXIS SAFE: pair positionally; works for 1, 2 or more crossed axes
        if axes:
            out[token] = axes[: len(row_tokens)] if row_tokens else axes
    return out


exp_inventory_ids = tuple(
    c["inventory_id"] for c in columns if c["inventory_id"] and not c["is_label_col"]
)
inv_meta = load_inventory_meta(client, exp_inventory_ids)

# Facets are project-scoped -> load per selected project and merge (counts of
# the same value are summed across projects).
_facets_acc: dict[str, dict[str, int]] = {}
for _plabel, _pid, _pshort in selected_projects:
    for _param, _vals in load_facets(client, _pid).items():
        _d = _facets_acc.setdefault(_param, {})
        for _nm, _cnt in _vals:
            _d[_nm] = _d.get(_nm, 0) + _cnt
facets = {
    p: sorted(d.items(), key=lambda kv: (-kv[1], kv[0].lower()))
    for p, d in _facets_acc.items()
}

# Property Tasks + DT index from EVERY selected project, in project order
# (one tasks.get_all call per project - see load_property_task_catalog).
_catalog = load_property_task_catalog(
    client, tuple(PROJECT_IDS), int(st.session_state.get("dt_cache_bust", 0))
)
dt_index: dict[str, dict] = _catalog["dt_index"]
_short_of_pid = {pid: short for _, pid, short in selected_projects}
property_tasks: list[dict] = []
for _t in _catalog["tasks"]:
    _t = dict(_t)
    _t["project"] = _short_of_pid.get(_t["project_id"], _t["project_id"])
    property_tasks.append(_t)

# ONE results store for the whole comparison, keyed by task id (globally
# unique). Tasks belonging to a project that was removed from the comparison
# are pruned so the tables and downloads never leak stale projects.
# NOTE: the DT-first rework was specced as `results_store::v4::`, but v4 was
# already taken by the task-first flow whose records lack dt_id/dc_id/unit_id
# - bumped to v5 so every stale cache is guaranteed to bust.
RESULTS_STORE_KEY = "results_store::v5"
_store = st.session_state.setdefault(RESULTS_STORE_KEY, {})
_valid_task_ids = {t["id"] for t in property_tasks}
for _tid in [k for k in _store if k not in _valid_task_ids]:
    del _store[_tid]

# --- PREDECESSOR: only lives in the worksheet's Apps design PDC row ----------
# It is NOT a field on InventoryItem (not top-level, not in Metadata, no facet).
# We read the Apps grid via the per-design /grid endpoint, which is NOT subject
# to the 20k-item truncation that hits sheets.get_cell_values on large sheets.
def _apps_row_values(row_type: str) -> dict[str, str]:
    """Union across every matching row - the merged multi-project worksheet can
    carry one PDC row per project when their names/paths differ."""
    sec = section_by_attr.get("app_design")
    if not sec:
        return {}
    out: dict[str, str] = {}
    for r in sec["rows"]:
        if r["type_raw"].split(".")[-1].upper() == row_type:
            for inv, v in r["values"].items():
                if v:
                    out.setdefault(inv, v)
    return out


predecessor_by_inv = _apps_row_values("PDC")
for _inv, _m in inv_meta.items():
    _m["predecessor"] = predecessor_by_inv.get(_inv, "")

# inventory_id -> set of data-template names (via the Property Tasks it appears in)
dt_ids_all = tuple({i for t in property_tasks for i in t["data_template_ids"]})
dt_name_of = load_data_template_names(client, dt_ids_all)
dts_of_inv: dict[str, set[str]] = {}
for t in property_tasks:
    names = {
        dt_name_of.get(i) or nm
        for i, nm in zip(t["data_template_ids"], t["data_templates"])
    }
    for inv in t["inventory_ids"]:
        dts_of_inv.setdefault(inv, set()).update(n for n in names if n)
all_data_templates = sorted({dt for s in dts_of_inv.values() for dt in s})


def column_header(c: dict) -> tuple[str, str]:
    """(top, bottom) header: top = short code (e.g. MO13137-053), bottom = name."""
    if not c["inventory_id"]:
        return (c["name"] or c["column_id"] or "", "")
    meta = inv_meta.get(c["inventory_id"], {})
    long_name = c["name"] or meta.get("name") or c["formulation_name"]
    candidates = [meta.get("alias", ""), meta.get("name", ""), c["formulation_name"]]
    code = ""
    for cand in candidates:
        if cand and cand != long_name and len(cand) <= 40:
            code = cand
            break
    if not code:
        code = _strip_inv(c["inventory_id"])
    # In a multi-project comparison the description carries the column's origin
    # (project / sheet), shown in header tooltips and the description row.
    origin = c.get("origin") or ""
    if origin:
        long_name = f"{long_name}  ·  {origin}" if long_name else origin
    return (code, long_name)


# ===========================================================================
# 4) GLOBAL FILTERS - one state, applied to every section table + downloads
#    Re-implementation of Albert's 7 client-side UI filters (README §9).
# ===========================================================================
st.header("2️⃣ Filters")

exp_cols_all = [c for c in columns if c["inventory_id"] and not c["is_label_col"]]

# Ingredient candidates = Product Design INV rows ("Contains inventory" filter)
product_section = section_by_attr.get("product_design")
inv_rows_product = (
    [r for r in product_section["rows"] if r["type_raw"].split(".")[-1] == "INV"]
    if product_section
    else []
)
ingredient_options = sorted({r["name"] for r in inv_rows_product if r["name"]})


def _with_none(options: list[str], has_blank: bool) -> list[str]:
    """Prepend the (None) sentinel when some column has no value for this attribute."""
    return ([NONE_LABEL] if has_blank else []) + options


def _options_from(attr: str, facet_param: str | None = None) -> list[str]:
    """Filter options = every value present on THIS sheet's columns, unioned with
    Albert's own facet list for the project (so nothing that exists is missing).
    Sorted by how many of the visible experiment columns carry it."""
    counts: dict[str, int] = {}
    for c in exp_cols_all:
        m = inv_meta.get(c["inventory_id"], {})
        vals = m.get(attr) or []
        if isinstance(vals, str):
            vals = [vals] if vals else []
        for v in vals:
            counts[v] = counts.get(v, 0) + 1
    if facet_param:
        for name, _cnt in facets.get(facet_param, []):
            counts.setdefault(name, 0)
    return sorted(counts, key=lambda v: (-counts[v], v.lower()))


all_tags = _with_none(
    _options_from("tags", "tags"),
    any(not m.get("tags") for m in inv_meta.values()),
)
all_creators = _with_none(
    _options_from("created_by", "createdBy"),
    any(not m.get("created_by") for m in inv_meta.values()),
)
all_predecessors = _with_none(
    _options_from("predecessor"),
    any(not m.get("predecessor") for m in inv_meta.values()),
)

f1, f2, f3 = st.columns(3)
with f1:
    # FIX #2: Albert's Formula/Product ID filter is a searchable dropdown, not a
    # free-text box. Streamlit's multiselect does substring type-ahead over the
    # option labels, so typing "MO13137-09" or "85p (PA" narrows the list live.
    exp_options = {
        f"{code}  ·  {desc}" if desc else code: c["inventory_id"]
        for c, (code, desc) in zip(
            exp_cols_all, [column_header(c) for c in exp_cols_all]
        )
    }
    flt_exp_labels = st.multiselect(
        "Formula / Product ID",
        list(exp_options.keys()),
        help="Type any part of the ID or the name (e.g. 'MO13137-09' or '85p (PA') "
        "to narrow the list, then pick the experiments you want.",
    )
    flt_exp_invs = {exp_options[l] for l in flt_exp_labels}
    flt_ingredients = st.multiselect(
        "Contains ingredient (Product Design)",
        ingredient_options,
        help="Show only formulations that have a value in the selected ingredient row(s).",
    )
with f2:
    flt_tags = st.multiselect(
        "Tags", all_tags, help=f"'{NONE_LABEL}' = formulations with no tag."
    )
    flt_creators = st.multiselect(
        "Created by", all_creators, help=f"'{NONE_LABEL}' = creator unknown."
    )
with f3:
    flt_preds = st.multiselect(
        "Predecessor",
        all_predecessors,
        help="Formulations derived from the selected predecessor formula(s). "
        f"'{NONE_LABEL}' = formulations with no predecessor.",
    )
    # The old "Data Templates (Results only)" multiselect moved into the
    # "Selection of Data Templates in Results" sub-section below, where it
    # drives WHAT the Results section downloads (DataTemplate-first flow).

f4, f5, f6, f7 = st.columns(4)
with f4:
    match_all = st.checkbox(
        "Match ALL conditions within a filter",
        value=False,
        help="e.g. must contain every selected ingredient / carry every selected tag.",
    )
with f5:
    flt_lock = st.radio("Lock state", ["All", "Locked", "Unlocked"], horizontal=True)
with f6:
    show_hidden = st.checkbox("Show hidden columns", value=False)
with f7:
    focus_view = st.checkbox(
        "Hide empty rows (Focus view)",
        value=False,
        help="Hide rows that have no value in any of the visible experiment columns.",
    )


# ===========================================================================
# 4b) ADVANCED FILTER - cascade conditions that hide formulation COLUMNS.
#     Each filter drills a section's row hierarchy to a target row (or a whole
#     group), attaches a condition (> / < / = / Range / exists / does not exist)
#     on that row's value, and keeps only the formulations that satisfy EVERY
#     applied filter (AND). Results filters need the Property Task loaded first.
# ===========================================================================
ADV_ANY = "(any)"


def _safe_selectbox(label: str, options: list[str], key: str, **kw):
    """selectbox whose stored value is reset when it falls out of `options`
    (cascade child options change when a parent changes - Streamlit would
    otherwise raise on the stale session_state value)."""
    if key in st.session_state and st.session_state[key] not in options:
        del st.session_state[key]
    return st.selectbox(label, options, key=key, **kw)


def _safe_multiselect(label: str, options: list[str], key: str, default=None, **kw):
    """multiselect whose stored values are pruned when they fall out of
    `options`; `default` seeds the state only on first render."""
    if key in st.session_state:
        st.session_state[key] = [v for v in st.session_state[key] if v in options]
    else:
        st.session_state[key] = [v for v in (default or []) if v in options]
    return st.multiselect(label, options, key=key, **kw)


def _popover(label: str, **kw):
    """st.popover (compact dropdown panel) with an expander fallback for older
    Streamlit versions."""
    if hasattr(st, "popover"):
        return st.popover(label, **kw)
    return st.expander(label)


def _bordered():
    """Bordered container (fallback to a plain one on older Streamlit)."""
    try:
        return st.container(border=True)
    except TypeError:  # Streamlit < 1.29
        return st.container()


# ===========================================================================
# CUSTOM ROW / GROUP ORDERING - shared by every section table + the XLSX.
#
# Two per-table stores in session state:
#   gorder::{table_key}  '¦'.join(node path) -> ordinal among its SIBLINGS
#   rorder::{table_key}  row_id              -> ordinal within its LEAF BLOCK
#
# A row's sort key is the ordinal of each of its ancestor nodes (compared only
# against true siblings - deeper components are reached only when every parent
# ordinal ties, i.e. within the same parent) followed by its own ordinal. So a
# row can NEVER leave its group/subgroup: reordering only permutes blocks among
# siblings and rows inside their own block.
# ===========================================================================
def _apply_custom_order(items: list[tuple[tuple, str, Any]], table_key: str) -> list:
    """items = [(path tuple, row_id, payload)] in original order. Returns the
    payloads re-sorted by the user's custom group/row ordering (stable: with no
    stored ordering the original order is preserved).

    Every 'unit' - a group/subgroup node OR a loose row sitting directly at
    that level - defaults to its FIRST-APPEARANCE item index, so node and row
    ordinals share one scale and reordering one never scrambles the other.
    Reorders always PERMUTE existing ordinal values (see the swap/move
    helpers), keeping that shared scale intact. The effective ordinals are
    published in session state ('ordmeta::' / 'ordmeta_rows::') for the
    reorder UI to read."""
    gorder = st.session_state.get(f"gorder::{table_key}") or {}
    rorder = st.session_state.get(f"rorder::{table_key}") or {}
    seen: dict[tuple, float] = {}
    for i, (path, _rid, _p) in enumerate(items):
        for k in range(1, len(path) + 1):
            seen.setdefault(path[:k], float(i))
    node_ord = {n: float(gorder.get("¦".join(n), d)) for n, d in seen.items()}
    row_ord = {
        rid: float(rorder.get(rid, i)) for i, (_path, rid, _p) in enumerate(items)
    }
    st.session_state[f"ordmeta::{table_key}"] = node_ord
    st.session_state[f"ordmeta_rows::{table_key}"] = row_ord

    def sort_key(pair):
        i, (path, rid, _p) = pair
        key = [node_ord[path[:k]] for k in range(1, len(path) + 1)]
        key.append(row_ord[rid])
        key.append(float(i))  # stability
        return key

    return [p for _, (_, _, p) in sorted(enumerate(items), key=sort_key)]


def _swap_group_order(table_key: str, node_a: tuple, node_b: tuple) -> None:
    """Swap two sibling nodes by EXCHANGING their effective ordinal values
    (value permutation keeps the shared node/row ordinal scale intact)."""
    meta = st.session_state.get(f"ordmeta::{table_key}") or {}
    va, vb = meta.get(node_a), meta.get(node_b)
    if va is None or vb is None:
        return
    g = st.session_state.setdefault(f"gorder::{table_key}", {})
    g["¦".join(node_a)], g["¦".join(node_b)] = float(vb), float(va)


def _section_order_items(section) -> list[tuple[tuple, str, dict]]:
    """(effective path, row_id, row) per section row. The EFFECTIVE path of a
    group/subgroup HEADER row includes its own name, so the header belongs to
    its own block and moves with it (e.g. the 'Neutralization degree (ND)' BLK
    row travels with the 'Neutralization degree (ND)' subgroup)."""
    rows = section["rows"]
    ancestors = {aid for r in rows for aid in (r.get("path_ids") or [])}
    items = []
    for r in rows:
        raw = str(r["row_id"]).split("::")[-1]  # merged ids are 'PID::ROWn'
        eff = tuple(r["path"]) + (
            (r["name"],) if raw in ancestors and r["name"] else ()
        )
        items.append((eff, str(r["row_id"]), r))
    return items


# ===========================================================================
# 4c) SELECTION OF DATA TEMPLATES IN RESULTS - DataTemplate-first flow.
#     Replaces the old "Data Templates (Results only)" filter. Each selector
#     row picks ONE Data Template, its Data Columns and (optionally) one
#     setpoint per interval axis. Everything here is METADATA ONLY (tasks /
#     data_templates / workflows) - property data is downloaded exclusively
#     by the "Load Data" button in the Results section.
# ===========================================================================
def dt_interval_axes(dt_id: str) -> list[dict]:
    """Interval setpoints per axis for ONE Data Template: the union across
    EVERY workflow of every block where the DT occurs (the same DT can sit on
    different workflows in different tasks). Resolution goes through the
    existing `_workflow_interval_map` (incl. its raw-HTTP fallback), and the
    axis count is DERIVED from the resolved intervalString - never hardcoded.
    Deduplication is by human-readable label. Returns
        [{"values": [label, ...], "names": [axis-parameter names seen]}, ...]"""
    wf_ids = sorted(
        {
            o["workflow_id"]
            for o in dt_index.get(dt_id, {}).get("occurrences", [])
            if o["workflow_id"]
        }
    )
    cache: dict[str, dict[str, list[str]]] = st.session_state.setdefault("wf_intervals", {})
    for w in wf_ids:
        if w not in cache:
            cache[w] = _workflow_interval_map(client, w)
    axes: list[dict] = []
    for w in wf_ids:
        for _token, labels in cache[w].items():
            for i, lab in enumerate(labels):
                while len(axes) <= i:
                    axes.append({"values": [], "names": []})
                if lab not in axes[i]["values"]:
                    axes[i]["values"].append(lab)
                nm = lab.split(":", 1)[0].strip() if ":" in lab else ""
                if nm and nm not in axes[i]["names"]:
                    axes[i]["names"].append(nm)
    return axes


st.markdown("**Selection of Data Templates in Results**")
st.caption(
    "Choose WHICH Data Templates the Results section will download - and, per "
    "template, which Data Columns and interval setpoints to show. This panel "
    "makes metadata calls only; property data is fetched by **Load Data** in "
    "the Results section."
)

_dt_label_of = {
    f"{v['name'] or dt}  ({dt})": dt
    for dt, v in sorted(
        dt_index.items(), key=lambda kv: ((kv[1]["name"] or "").lower(), kv[0])
    )
}
DT_PLACEHOLDER = "(select a Data Template)"


# --- "Select all Data Templates with values" (one-click auto-selection) ------
# Click -> (1) ONE lightweight `check_for_task_data` call per Property Task
# (parallel; it returns block x inventory combos with a `data_exists` flag and
# downloads ZERO property values) finds the Data Templates that actually
# contain data and fills one selector row per hit; (2) Load Data is triggered
# automatically for those templates (same fetch the button below would do);
# (3) once the records are in the store, each row's Data Columns and interval
# setpoints are pruned to the ones that actually carry values (deferred prune,
# see `dtsel_autoprune_pending` below).
def _dts_with_data() -> set[str]:
    from concurrent.futures import ThreadPoolExecutor

    dt_of_block: dict[tuple[str, str], set[str]] = {}
    for _dt, _ent in dt_index.items():
        for _o in _ent["occurrences"]:
            dt_of_block.setdefault((_o["task_id"], _o["block_id"]), set()).add(_dt)
    tids = sorted({t for t, _b in dt_of_block})
    if not tids:
        return set()

    def _check(tid: str):
        try:
            return tid, list(client.property_data.check_for_task_data(task_id=tid)), None
        except Exception as e:  # noqa: BLE001
            return tid, [], f"{type(e).__name__}: {e}"

    out: set[str] = set()
    errs: list[str] = []
    workers = int(st.session_state.get("fetch_workers", 16))
    with ThreadPoolExecutor(max_workers=min(workers, len(tids))) as ex:
        for tid, combos, err in ex.map(_check, tids):
            if err:
                errs.append(f"{tid}: {err}")
                continue
            for c in combos:
                # missing flag -> assume data may exist (same default as Load
                # Data's planner); the deferred prune corrects any over-selection
                if not getattr(c, "data_exists", True):
                    continue
                for _dt in dt_of_block.get((tid, getattr(c, "block_id", "") or ""), ()):
                    out.add(_dt)
    if errs:
        st.warning(f"Data check failed for {len(errs)} task(s) - e.g. {errs[0]}")
    return out


_autosel_clicked = st.button(
    "⚡ Select all Data Templates with values",
    key="dtsel_autoselect",
    disabled=not _dt_label_of,
    help="One click: finds every Data Template that actually contains property "
    "data (one lightweight data-exists check per task - no values downloaded), "
    "fills one selector row per hit, downloads their data (same as Load Data) "
    "and then unselects the Data Columns / interval setpoints that turned out "
    "to hold no values. Replaces the rows configured below.",
)
if _autosel_clicked:
    with st.spinner("Checking which Data Templates contain data (no values downloaded)..."):
        _dts_hit = _dts_with_data()
    for _k in [k for k in st.session_state if str(k).startswith("dtsel::")]:
        del st.session_state[_k]
    _labels_hit = [_l for _l, _d in _dt_label_of.items() if _d in _dts_hit]
    st.session_state["dtsel_count"] = len(_labels_hit)
    for _j, _l in enumerate(_labels_hit):
        st.session_state[f"dtsel::{_j}::dt"] = _l
    if _labels_hit:
        st.session_state["dtsel_autoload_pending"] = True
        st.session_state["dtsel_autoprune_pending"] = True
    else:
        st.info("No Data Template in the selected project(s) contains property data.")

# --- deferred prune: runs on the first render where EVERY task behind the
# selected templates is in the results store (i.e. right after the
# auto-triggered Load Data finished, or immediately when it was all cached).
# Computes, per Data Template, which Data Columns and which resolved interval
# setpoints actually carry values; the row loop below writes that into the
# widgets' session state BEFORE they are instantiated.
_dtsel_autoprune = False
_dt_usage: dict[str, dict] = {}
if st.session_state.get("dtsel_autoprune_pending"):
    _sel_ids_now = {
        _dt_label_of[st.session_state.get(f"dtsel::{_j}::dt")]
        for _j in range(int(st.session_state.get("dtsel_count", 0)))
        if st.session_state.get(f"dtsel::{_j}::dt") in _dt_label_of
    }
    _needed_tids = {
        _o["task_id"]
        for _d in _sel_ids_now
        for _o in dt_index.get(_d, {}).get("occurrences", [])
    }
    _store_now = st.session_state.get(RESULTS_STORE_KEY, {})
    if _sel_ids_now and _needed_tids <= set(_store_now):
        st.session_state["dtsel_autoprune_pending"] = False
        _dtsel_autoprune = True
        _prune_recs = [
            _r
            for _tid in _needed_tids
            for _r in _store_now.get(_tid, [])
            if "__error__" not in _r
        ]
        # resolve interval tokens with the cached workflow maps (fetch missing)
        _wf_cache = st.session_state.setdefault("wf_intervals", {})
        _missing_wids = sorted(
            {
                _w
                for _r in _prune_recs
                for _w in ([_r.get("workflow_id", "")] + list(_r.get("task_workflows") or []))
                if _w and _w not in _wf_cache
            }
        )
        if _missing_wids:
            from concurrent.futures import ThreadPoolExecutor

            with ThreadPoolExecutor(max_workers=min(8, len(_missing_wids))) as _ex:
                for _w, _m in zip(
                    _missing_wids,
                    _ex.map(lambda w: _workflow_interval_map(client, w), _missing_wids),
                ):
                    _wf_cache[_w] = _m
        for _r in _prune_recs:
            _d = _r.get("dt_id") or ""
            if not _d:
                continue
            _u = _dt_usage.setdefault(
                _d, {"dcs": set(), "iv": [set(), set()], "unresolved": False}
            )
            if _r.get("dc_id"):
                _u["dcs"].add(_r["dc_id"])
            _raw = str(_r.get("raw_interval", "") or "")
            if _raw:
                _labels: list[str] = []
                for _w in [_r.get("workflow_id", "")] + list(_r.get("task_workflows") or []):
                    if _w and _wf_cache.get(_w, {}).get(_raw):
                        _labels = _wf_cache[_w][_raw]
                        break
                if not _labels:
                    _u["unresolved"] = True
                for _ax_j in (0, 1):
                    if _ax_j < len(_labels) and _labels[_ax_j]:
                        _u["iv"][_ax_j].add(_labels[_ax_j])

_dtsel_count = int(st.session_state.get("dtsel_count", 0))
if _dtsel_count == 0:
    if st.button("➕ Add DT", key="dtsel_add_first", disabled=not _dt_label_of):
        st.session_state["dtsel_count"] = 1
        st.rerun()
if not _dt_label_of:
    st.caption("No Property Tasks with Data Templates found in the selected project(s).")

dt_selectors: list[dict] = []
for _i in range(_dtsel_count):
    _cols = st.columns([1.7, 2.0, 1.4, 1.4])
    with _cols[0]:
        _lbl = _safe_selectbox(
            "Data Templates",
            [DT_PLACEHOLDER] + list(_dt_label_of),
            key=f"dtsel::{_i}::dt",
            help="Every Data Template that occurs on a Property Task of the "
            "selected project(s), from the task index (no property data loaded).",
        )
    if _lbl == DT_PLACEHOLDER:
        with _cols[1]:
            st.multiselect(
                "Data Columns", [], key=f"dtsel::{_i}::dc_ph", disabled=True
            )
        with _cols[2]:
            st.multiselect("Interval 1", [], key=f"dtsel::{_i}::iv1_ph", disabled=True)
        with _cols[3]:
            st.multiselect("Interval 2", [], key=f"dtsel::{_i}::iv2_ph", disabled=True)
        continue
    _dt_id = _dt_label_of[_lbl]

    # changing the DT of a row resets its dependent Data Columns / Intervals
    if st.session_state.get(f"dtsel::{_i}::dt_prev") != _dt_id:
        for _sfx in ("dc", "iv1", "iv2"):
            st.session_state.pop(f"dtsel::{_i}::{_sfx}", None)
        st.session_state[f"dtsel::{_i}::dt_prev"] = _dt_id

    _dcs = load_dt_definition(client, _dt_id, int(st.session_state.get("dt_cache_bust", 0)))
    _hidden_of = {_d["id"]: _d["hidden"] for _d in _dcs}
    _dc_label_of = {
        f"{_d['name']}  ({_d['id']})"
        + (f"  [{_d['unit_symbol']}]" if _d["unit_symbol"] else ""): _d["id"]
        for _d in _dcs
    }
    _default_dcs = [_l for _l, _did in _dc_label_of.items() if not _hidden_of.get(_did)]
    # deferred auto-prune: keep only the Data Columns that actually carry values
    # (a DT with zero usable records keeps its defaults - never an empty match-all)
    if _dtsel_autoprune and _dt_id in _dt_usage:
        st.session_state[f"dtsel::{_i}::dc"] = [
            _l for _l, _did in _dc_label_of.items() if _did in _dt_usage[_dt_id]["dcs"]
        ]
    with _cols[1]:
        _sel_dc_labels = _safe_multiselect(
            "Data Columns",
            list(_dc_label_of),
            key=f"dtsel::{_i}::dc",
            default=_default_dcs,
            help="From the Data Template definition (the unit symbols shown "
            "here also label the Results, since property data carries only "
            "unit ids). Defaults to every non-hidden data column.",
        )

    # MULTI-SELECT per interval axis. Each axis only ever offers the setpoints
    # of the parameter the workflow assigns to that axis position (axis 1 =
    # first segment of intervalString, e.g. Time; axis 2 = second, e.g. Speed)
    # - so several setpoints of the SAME parameter can be combined, and mixing
    # parameters across axes is impossible by construction. An EMPTY selection
    # means '(any)': the axis is not filtered.
    _axes = dt_interval_axes(_dt_id)
    _iv_sel: list[list[str]] = []
    for _ax_i in (0, 1):
        _ax = _axes[_ax_i] if _ax_i < len(_axes) else None
        _has = bool(_ax and _ax["values"])
        _iv_label = (
            f"Interval {_ax_i + 1} — {_ax['names'][0]}"
            if _has and len(_ax["names"]) == 1
            else f"Interval {_ax_i + 1}"
        )
        # deferred auto-prune: select only the setpoints that carry values.
        # Unresolved tokens or full coverage fall back to an EMPTY selection
        # ('(any)' - no filter), so pruning can never drop a real record.
        if _dtsel_autoprune and _dt_id in _dt_usage and _has:
            _u = _dt_usage[_dt_id]
            _keep = [_v2 for _v2 in _ax["values"] if _v2 in _u["iv"][_ax_i]]
            if _u["unresolved"] or len(_keep) == len(_ax["values"]):
                _keep = []
            st.session_state[f"dtsel::{_i}::iv{_ax_i + 1}"] = _keep
        with _cols[2 + _ax_i]:
            _v = _safe_multiselect(
                _iv_label,
                _ax["values"] if _has else [],
                key=f"dtsel::{_i}::iv{_ax_i + 1}",
                disabled=not _has,
                placeholder=ADV_ANY,
                help=(
                    "Pick ANY NUMBER of setpoints on this axis (unioned across "
                    "every workflow this Data Template occurs on). Leave it "
                    f"empty for '{ADV_ANY}' - no filter on this axis."
                    if _has
                    else "This Data Template has no interval on this axis."
                ),
            )
            if _has and len(_ax["names"]) > 1:
                st.caption(
                    f"⚠️ Axis {_ax_i + 1} parameter differs across workflows: "
                    + ", ".join(_ax["names"])
                )
        _iv_sel.append(_v)

    _occ = dt_index.get(_dt_id, {}).get("occurrences", [])
    dt_selectors.append(
        {
            "dt_id": _dt_id,
            "dt_label": _lbl,
            "dt_name": dt_index.get(_dt_id, {}).get("name", "") or _dt_id,
            "dc_ids": {_dc_label_of[_l] for _l in _sel_dc_labels},
            "iv1": _iv_sel[0],
            "iv2": _iv_sel[1],
            # match on (task_id, block_id) - block ids repeat across tasks
            "occ": {(o["task_id"], o["block_id"]) for o in _occ},
            "unit_symbol_of": {_d["id"]: _d["unit_symbol"] for _d in _dcs},
            "axes": _axes,
        }
    )

if _dtsel_count > 0:
    _db = st.columns([1.2, 1.5, 4])
    with _db[0]:
        if st.button("➕ Add new DT", key="dtsel_add_more"):
            st.session_state["dtsel_count"] = _dtsel_count + 1
            st.rerun()
    with _db[1]:
        if st.button(
            "🗑️ Remove last DT", key="dtsel_remove", disabled=_dtsel_count <= 1
        ):
            _j = _dtsel_count - 1
            for _k in [k for k in st.session_state if str(k).startswith(f"dtsel::{_j}::")]:
                del st.session_state[_k]
            st.session_state["dtsel_count"] = _dtsel_count - 1
            st.rerun()

# read by the Load Data button and by the Results record filter
st.session_state["dt_selectors"] = dt_selectors


# ===========================================================================
# 4c-bis) RESULTS LOAD CONTROLS - these used to sit above the Results table;
#     they now live HERE, between "Selection of Data Templates in Results"
#     and the Advanced filter. Only the CONTROLS moved: the actual fetch (the
#     progress bar) and the Results table itself still happen in the Results
#     section below, on the same run.
# ===========================================================================
include_foreign = False
long_view = False
_load_clicked = False
target_tasks: dict[str, dict] = {}
_results_to_fetch: list[dict] = []
loaded_recs: list[dict] = []
RESULTS_PROGRESS_SLOT = None  # placeholder rendered under the Load Data button
results_store = st.session_state.setdefault(RESULTS_STORE_KEY, {})

if section_by_attr.get("result_design"):
    _rc1, _rc2 = st.columns(2)
    with _rc1:
        include_foreign = st.checkbox(
            "Include experiments filtered out / from other sheets",
            value=False,
            help="A Property Task can hold data for experiments hidden by the "
            "filters above or living on another sheet.",
        )
    with _rc2:
        long_view = st.checkbox("Long (tidy) view instead of pivot", value=False)

    agg_choice = st.radio(
        "Repeated measurements per property",
        ["List all values (6.12 | 6.65 | 5.71)", "Average"],
        horizontal=True,
        key="agg::results",
        help="When one property has several measurements for the same experiment, "
        "either list every value or show their numeric average. Applies to the pivot "
        "view on screen and to the XLSX / CSV (pivot) downloads.",
    )
    st.session_state["results_agg_mode"] = "avg" if agg_choice == "Average" else "list"

    # --- Load plan: union of the tasks behind every configured DT row ---------
    for _sel in dt_selectors:
        for _o in dt_index.get(_sel["dt_id"], {}).get("occurrences", []):
            target_tasks.setdefault(
                _o["task_id"], {"id": _o["task_id"], "name": _o["task_name"]}
            )
    _results_to_fetch = [t for tid, t in target_tasks.items() if tid not in results_store]

    if not dt_selectors:
        st.info(
            "Configure at least one Data Template in **Selection of Data "
            "Templates in Results** above - then press Load Data."
        )

    _lc1, _lc2, _lc3 = st.columns([1.5, 1.6, 1.2])
    with _lc1:
        _load_clicked = st.button(
            f"⬇️ Load Data ({len(target_tasks)} task(s))",
            type="primary",
            disabled=not target_tasks,
            help="Downloads property data ONLY for the tasks where the selected "
            "Data Templates occur (already-loaded tasks are reused from cache). "
            "Data Column / interval choices filter the display and cost nothing.",
        )
    with _lc2:
        _reload_clicked = st.button(
            "🔄 Reload (discard caches)",
            disabled=not target_tasks,
            help="Busts the DT/task index, the Data Template definitions, the "
            "workflow interval cache and the results store, then re-fetches.",
        )
    with _lc3:
        st.session_state["fetch_workers"] = st.number_input(
            "Parallel requests",
            min_value=1,
            max_value=48,
            value=int(st.session_state.get("fetch_workers", 16)),
            step=4,
            help="Size of the flat request pool. Load Data issues one request per "
            "block x inventory combo of the SELECTED Data Templates, across all "
            "target tasks at once. Raising this shortens the load; back off on errors.",
        )

    # "Select all Data Templates with values" also presses Load Data itself:
    # the flag is consumed exactly once; with everything already cached there
    # is nothing to fetch and the prune above has already run on this render.
    if st.session_state.get("dtsel_autoload_pending"):
        st.session_state["dtsel_autoload_pending"] = False
        if _results_to_fetch:
            _load_clicked = True

    # The download progress bar renders HERE, just below the Load Data button
    # (the fetch itself still runs from the Results section, into this slot).
    RESULTS_PROGRESS_SLOT = st.container()

    if _reload_clicked:
        st.session_state["dt_cache_bust"] = int(st.session_state.get("dt_cache_bust", 0)) + 1
        load_property_task_catalog.clear()
        load_dt_definition.clear()
        st.session_state["wf_intervals"] = {}
        st.session_state["wf_unresolved"] = {}
        st.session_state[RESULTS_STORE_KEY] = {}
        st.rerun()

    _loaded_n_ctrl = sum(1 for _tid in target_tasks if _tid in results_store)
    if dt_selectors and not _loaded_n_ctrl:
        st.caption("No property data loaded yet - press **Load Data**.")

    # --- Diagnostics: how the DT-first index maps to the load plan ------------
    with st.expander("🔧 DT index & Load plan (DataTemplate-first)"):
        st.write("**dt_index** (from `tasks.get_all` inline Blocks - zero property data):")
        st.json(dt_index, expanded=False)
        if dt_selectors:
            st.write("**Resolved interval axes per configured DT:**")
            st.write(
                {
                    f"{_sel['dt_name']} ({_sel['dt_id']})": [
                        {"axis": i + 1, "parameters": ax["names"], "values": ax["values"]}
                        for i, ax in enumerate(_sel["axes"])
                    ]
                    or "(no intervals on any workflow)"
                    for _sel in dt_selectors
                }
            )
        st.write(
            "**Tasks that Load Data will fetch** (union across DT rows, deduped):",
            sorted(target_tasks),
        )


def _adv_row_value(row: dict, field: str) -> str:
    """Value of a cascade field (Name / Group / Subgroup n) for a Product/Process row."""
    if field == "Name":
        return row.get("name", "") or ""
    path = row.get("path", []) or []
    if field == "Group":
        return path[0] if len(path) > 0 else ""
    if field.startswith("Subgroup "):
        try:
            i = int(field.split()[1])
        except ValueError:
            return ""
        return path[i] if len(path) > i else ""
    return ""


# Cascade fields each filterable section exposes (Apps excluded).
_adv_section_fields: dict[str, list[str]] = {}
if section_by_attr.get("product_design"):
    _pd_depth = section_by_attr["product_design"]["max_depth"]
    _adv_section_fields["product_design"] = (
        ["Group"] + [f"Subgroup {i}" for i in range(1, max(_pd_depth, 1))] + ["Name"]
    )
if section_by_attr.get("result_design"):
    _adv_section_fields["result_design"] = [
        "Data Template", "Data Column", "Interval 1", "Interval 2"
    ]
if section_by_attr.get("process_design"):
    _adv_section_fields["process_design"] = ["Name", "Group"]

# Loaded Results records (populated lazily lower on the page; persist across reruns).
_adv_result_records = [
    r
    for recs in st.session_state.get(RESULTS_STORE_KEY, {}).values()
    for r in recs
    if isinstance(r, dict) and "__error__" not in r
]

# The 'Results' entry of the Section dropdown is shown as INACTIVE until the
# Results table has actually been downloaded (Load Data pressed): a native
# selectbox cannot grey out one option, so the state is written into the
# label itself. Once data arrives the label flips back to plain 'Results'
# (and _safe_selectbox resets any stale choice of the inactive label).
_results_loaded = bool(_adv_result_records)
_RESULTS_INACTIVE_LABEL = "Results  (inactive - press Load Data first)"
_adv_labels = {
    "product_design": "Product Design",
    "result_design": "Results" if _results_loaded else _RESULTS_INACTIVE_LABEL,
    "process_design": "Process Design",
}
_adv_label_to_attr = {
    _adv_labels[a]: a for a in _adv_labels if a in _adv_section_fields
}

adv_specs: list[dict] = []

if _adv_section_fields:
    st.markdown("**Advanced filter**")
    st.caption(
        "Build a condition on a row's value (an ingredient amount, a property "
        "result, ...) to keep only the formulations that satisfy it - across every "
        "table. Multiple filters are combined with AND. A Results filter applies "
        "to the ALREADY-DOWNLOADED results (press Load Data in the Results "
        "section first): it is a DISPLAY filter only and does not reduce what "
        "Load Data fetches from Albert."
    )

    adv_count = int(st.session_state.get("adv_filter_count", 0))
    if adv_count == 0:
        if st.button("➕ Add filter", key="adv_add_first"):
            st.session_state["adv_filter_count"] = 1
            st.rerun()

    _adv_field_slots = max((len(f) for f in _adv_section_fields.values()), default=1)
    _attr_options = list(_adv_label_to_attr.keys())

    for i in range(adv_count):
        st.markdown(f"**Filter {i + 1}**")
        seccol = st.columns([1.3] + [1.2] * _adv_field_slots)
        with seccol[0]:
            sec_label = _safe_selectbox("Section", _attr_options, key=f"advf::{i}::section")
        attr = _adv_label_to_attr[sec_label]
        fields = _adv_section_fields[attr]
        is_result = attr == "result_design"
        base = _adv_result_records if is_result else section_by_attr[attr]["rows"]

        # --- cascade -----------------------------------------------------------
        # Every level is shown; (any) means "don't constrain here" and the cascade
        # CONTINUES, so empty intermediate Subgroups can be left at (any) and you can
        # still drill down to Name. Each level's options come from the rows/records
        # that already match the non-(any) selections above it.
        selected: dict[str, str] = {}
        matching = list(base)
        for k, field in enumerate(fields):
            if is_result:
                opts = sorted(
                    {str(r.get(field, "")) for r in matching if str(r.get(field, "")).strip()}
                )
            else:
                opts = sorted(
                    {_adv_row_value(r, field) for r in matching if _adv_row_value(r, field)}
                )
            with seccol[min(k + 1, len(seccol) - 1)]:
                choice = _safe_selectbox(field, [ADV_ANY] + opts, key=f"advf::{i}::{attr}::{field}")
            if choice != ADV_ANY:
                selected[field] = choice
                if is_result:
                    matching = [r for r in matching if str(r.get(field, "")) == choice]
                else:
                    matching = [r for r in matching if _adv_row_value(r, field) == choice]

        # --- numeric target? decides which Logic operators are offered -------
        sample: list[str] = []
        if is_result:
            for r in matching:
                v = str(r.get("value", ""))
                if v:
                    sample.append(v)
                if len(sample) >= 300:
                    break
        else:
            for r in matching:
                for v in (r.get("values", {}) or {}).values():
                    if v:
                        sample.append(v)
                if len(sample) >= 300:
                    break
        numeric = bool(sample) and all(_to_number(v)[1] for v in sample)
        logic_opts = (
            [">", "<", "=", "Range", "exists", "does not exist"]
            if numeric
            else ["exists", "does not exist"]
        )

        # --- Logic + Value + Apply ------------------------------------------
        lvcol = st.columns([1.1, 1.1, 1.1, 1.0])
        with lvcol[0]:
            logic = _safe_selectbox("Logic", logic_opts, key=f"advf::{i}::logic")
        a = b = ""
        needs_value = logic in (">", "<", "=", "Range")
        if needs_value:
            with lvcol[1]:
                a = st.text_input("From" if logic == "Range" else "Value", key=f"advf::{i}::a")
            if logic == "Range":
                with lvcol[2]:
                    b = st.text_input("To", key=f"advf::{i}::b")
        with lvcol[3]:
            st.write("")  # nudge the button down to align with the inputs
            if st.button("✅ Apply filter", key=f"advf::{i}::applybtn"):
                st.session_state[f"advf::{i}::active"] = True
                st.rerun()

        # --- status + spec ---------------------------------------------------
        active = bool(st.session_state.get(f"advf::{i}::active"))
        if logic == "Range":
            complete = a.strip() != "" or b.strip() != ""
        elif needs_value:
            complete = a.strip() != ""
        else:
            complete = True
        available = len(matching) > 0 and (not is_result or len(_adv_result_records) > 0)

        crumb = " › ".join(selected.get(f, ADV_ANY) for f in fields) or ADV_ANY
        val_txt = (f" {a}" + (f"..{b}" if logic == "Range" else "")) if needs_value else ""
        summary = f"{sec_label} · {crumb} · {logic}{val_txt}"

        if not active:
            st.caption(f"Filter {i + 1} (not applied yet): {summary}")
        elif is_result and not available:
            st.info(
                f"Filter {i + 1}: press Load Data in the Results section below "
                f"to activate — {summary}"
            )
        elif not available:
            st.warning(f"Filter {i + 1}: no rows match the selection — {summary}")
        elif not complete:
            st.warning(f"Filter {i + 1}: enter a value to activate — {summary}")
        else:
            st.success(f"✓ Filter {i + 1}: {summary}")
            adv_specs.append(
                {"is_result": is_result, "matching": matching, "logic": logic, "a": a, "b": b}
            )

    if adv_count > 0:
        bcol = st.columns([1.3, 1.5, 4])
        with bcol[0]:
            if st.button("➕ Add new filter", key="adv_add_more"):
                st.session_state["adv_filter_count"] = adv_count + 1
                st.rerun()
        with bcol[1]:
            if st.button("🗑️ Remove last filter", key="adv_remove"):
                _j = adv_count - 1
                for _k in [k for k in st.session_state if str(k).startswith(f"advf::{_j}::")]:
                    del st.session_state[_k]
                st.session_state["adv_filter_count"] = max(0, adv_count - 1)
                st.rerun()


def _adv_spec_passes(spec: dict, inv_id: str) -> bool:
    logic, a, b = spec["logic"], spec["a"], spec["b"]
    if spec["is_result"]:
        vals = [str(r.get("value", "")) for r in spec["matching"] if r.get("inventory_id") == inv_id]
    else:
        vals = [row["values"].get(inv_id, "") for row in spec["matching"]]
    nonempty = [v for v in vals if str(v) != ""]
    if logic == "exists":
        return len(nonempty) > 0
    if logic == "does not exist":
        return len(nonempty) == 0
    # Precise comparison: the formulation passes only if it HAS a matching value
    # AND *every* matching value satisfies the condition. Using all() (not any())
    # means a formulation with any measurement that violates the threshold - e.g. a
    # Cobb Value of 10.2 under a "< 10" filter, be it a replicate trial or a value
    # from an interval the cascade did not fully pin down - is correctly excluded,
    # so no value shown in the table can contradict the applied filter.
    return len(nonempty) > 0 and all(_cmp_pass(v, logic, a, b) for v in nonempty)


def advanced_filter_passes(inv_id: str) -> bool:
    """AND of every applied Advanced filter. True when none are applied."""
    return all(_adv_spec_passes(s, inv_id) for s in adv_specs)


def _ingredient_hit(col: dict, wanted: list[str]) -> bool:
    inv = col["inventory_id"]
    hits = [
        any(r["name"] == w and r["values"].get(inv, "") != "" for r in inv_rows_product)
        for w in wanted
    ]
    return all(hits) if match_all else any(hits)


def _set_filter_hit(have: set[str], wanted: list[str]) -> bool:
    """(None) matches items with an empty set; real selections use the ANY/ALL
    toggle. The two are OR'd, so selecting every option shows every column."""
    real = [w for w in wanted if w != NONE_LABEL]
    if not have:
        return NONE_LABEL in wanted
    if not real:
        return False  # only (None) was selected, and this item has values
    return have.issuperset(real) if match_all else bool(have & set(real))


def _scalar_filter_hit(have: str, wanted: list[str]) -> bool:
    if not have:
        return NONE_LABEL in wanted
    return have in wanted


def column_passes(c: dict) -> bool:
    if not show_hidden and c["hidden"]:
        return False
    if flt_lock == "Locked" and not c["locked"]:
        return False
    if flt_lock == "Unlocked" and c["locked"]:
        return False

    meta = inv_meta.get(c["inventory_id"], {})

    if flt_exp_invs and c["inventory_id"] not in flt_exp_invs:
        return False

    if flt_ingredients and not _ingredient_hit(c, flt_ingredients):
        return False

    if flt_tags and not _set_filter_hit(set(meta.get("tags", [])), flt_tags):
        return False

    if flt_creators and not _scalar_filter_hit(meta.get("created_by", ""), flt_creators):
        return False

    if flt_preds and not _scalar_filter_hit(meta.get("predecessor", ""), flt_preds):
        return False

    return True


visible_cols = [
    c
    for c in exp_cols_all
    if column_passes(c) and advanced_filter_passes(c["inventory_id"])
]

n_hidden = sum(1 for c in exp_cols_all if c["hidden"])
st.caption(
    f"**{len(visible_cols)} / {len(exp_cols_all)} experiment columns** pass the filters"
    + (f" ({n_hidden} hidden in Albert)" if not show_hidden and n_hidden else "")
)
if not visible_cols:
    # An Advanced filter that no formulation satisfies (e.g. a value nobody hits,
    # or a Results filter whose Property block belongs to formulations that are not
    # columns on this sheet) would otherwise remove EVERY experiment column. A hard
    # st.stop() here blanks the whole Worksheet below - Product Design, Results and
    # all - so the Results table "disappears" the moment such a filter is applied.
    # When the emptiness is caused by the Advanced filter, warn loudly but keep the
    # page rendering (the section tables fall back to their key columns) so the user
    # can see what happened and relax the filter. A base-filter wipeout keeps the
    # original hard stop.
    if adv_specs:
        st.warning(
            "⚠️ **No experiment column passes the Advanced filter.** Every "
            "formulation was filtered out, so the tables below show their key "
            "columns only. Relax or remove the Advanced filter (or tick *Include "
            "experiments filtered out* in Results) to bring the columns back."
        )
    else:
        st.warning("No experiment column passes the current filters.")
        st.stop()

col_tuples = []
_seen_codes: dict[str, int] = {}
for _c in visible_cols:
    _code, _desc = column_header(_c)
    _code = _code or _strip_inv(_c["inventory_id"])
    if _code in _seen_codes:  # codes must be unique - they are the column labels
        _seen_codes[_code] += 1
        _code = f"{_code} ({_seen_codes[_code]})"
    else:
        _seen_codes[_code] = 1
    col_tuples.append((_code, _desc))

colid_to_tuple = dict(zip([c["column_id"] for c in visible_cols], col_tuples))
invid_to_tuple = {
    c["inventory_id"]: t for c, t in zip(visible_cols, col_tuples) if c["inventory_id"]
}


# ===========================================================================
# 5) Row display options + section tables
# ===========================================================================
st.header("3️⃣ Worksheet")

# Group / Subgroup columns are always built now; they can be hidden per-table via
# each table's "Hide columns" dropdown instead of a single global toggle.
show_hier_cols = True

o1, o2, o3, o4 = st.columns(4)
with o1:
    dec_choice = st.selectbox(
        "Decimals",
        ["All"] + list(range(0, 7)),
        index=0,
        help="Round numbers to this many significant decimals, counted from the first "
        "non-zero decimal place. e.g. Decimals=2 turns 0.0012345 -> 0.0012, "
        "1.0123 -> 1.012 and 1.123 -> 1.12. 'All' leaves numbers untouched.",
    )
    DECIMALS = None if dec_choice == "All" else int(dec_choice)
with o2:
    show_type_col = st.checkbox("Show 'Row type' column", value=False)
with o3:
    hide_blk = st.checkbox("Hide Blank (BLK) rows", value=False)
with o4:
    indent_names = st.checkbox(
        "Indent row names by depth", value=False, help="Mimics the Albert UI tree."
    )
    show_desc_row = st.checkbox(
        "Show experiment description row",
        value=True,
        help="Adds the full formulation name as the first row. The name is also "
        "available as a tooltip on each column header.",
    )

# --- 'Display' menu: reproduces the worksheet's Display dropdown -------------
# Compact popover (like Albert's Display button) that lists the HEADER NAME of
# every non-experiment sheet column (Blank / Lookup / Function / Result, ...)
# from the complete Formulas[] catalog - the type does not matter for the
# menu. Default selection = the columns whose Display toggle is ON in Albert
# (hidden == False). Selected columns are placed LEFT of the experiment
# columns, right after the key (Group / Subgroup / Name) columns.
DISPLAY_COL_LABELS = {
    "BLK": "Blank column",
    "LKP": "Lookup column",
    "FNC": "Function column",
    "RSL": "Result column",
    "FOR": "Formula column",
}
special_cols_all = [
    c for c in columns if (c.get("col_key") or "").startswith("COL::")
]
# unique display name per column (two columns may share a header)
_sp_seen: dict[str, int] = {}
for _c in special_cols_all:
    _nm = _c["name"] or _c["column_id"] or "column"
    if _nm in _sp_seen:
        _sp_seen[_nm] += 1
        _c["_disp_name"] = f"{_nm} ({_sp_seen[_nm]})"
    else:
        _sp_seen[_nm] = 1
        _c["_disp_name"] = _nm

_disp_col, _ = st.columns([1.2, 2.8])  # keep the menu small, not full-width
with _disp_col:
    if special_cols_all:
        with _popover("🗂️ Display"):
            st.caption(
                "Sheet columns (Blank / Lookup / Function / ...). Pick the ones "
                "to show - they appear right after the last Subgroup column, "
                "before the experiments. Default = the columns visible in "
                "Albert's own Display menu."
            )
            display_sel = _safe_multiselect(
                "Columns",
                [c["_disp_name"] for c in special_cols_all],
                key="display_cols",
                default=[c["_disp_name"] for c in special_cols_all if not c["hidden"]],
                help="The list shows every sheet column found on the worksheet "
                "grid, whatever its type.",
            )
    else:
        display_sel = []
        st.caption("Display: no Blank / Lookup / Function columns on this sheet.")
visible_special_cols = [
    c for c in special_cols_all if c.get("_disp_name") in (display_sel or [])
]

special_tuples: list[tuple[str, str]] = []
for _c in visible_special_cols:
    _code = _c["_disp_name"]
    if _code in _seen_codes:  # shares the uniqueness pool with the experiment codes
        _seen_codes[_code] += 1
        _code = f"{_code} ({_seen_codes[_code]})"
    else:
        _seen_codes[_code] = 1
    _lab = DISPLAY_COL_LABELS.get(str(_c["type"]).split(".")[-1], "Sheet column")
    _origin = _c.get("origin") or ""
    special_tuples.append((_code, _lab + (f"  ·  {_origin}" if _origin else "")))


def hier_cols_for(section: dict) -> list[str]:
    """['Group', 'Subgroup 1', ..., 'Subgroup n'] sized to this section's tree."""
    if not show_hier_cols or not section["max_depth"]:
        return []
    d = section["max_depth"]
    return ["Group"] + [f"Subgroup {i}" for i in range(1, d)]


def key_cols_for(section: dict) -> list[str]:
    return (
        ["Name"]
        + hier_cols_for(section)
        + (["Row type"] if show_type_col else [])
    )


def _row_in_filter(r: dict, row_filter: dict[int, list[str]]) -> bool:
    """row_filter: {ancestor_level -> allowed names}. Empty selection at a level
    = no filtering at that level. NONE_LABEL matches rows that have no ancestor
    at that level (i.e. the cell is blank), so 'select everything' really does
    mean everything. A group header row is kept when its own name is selected."""
    for level, wanted in row_filter.items():
        if not wanted:
            continue
        at_level = r["path"][level] if len(r["path"]) > level else ""
        # a header row sits AT this level: its own name is the value shown below it
        is_the_header = len(r["path"]) == level and r["name"] in wanted
        if at_level == "":
            if NONE_LABEL in wanted or is_the_header:
                continue
            return False
        if at_level not in wanted:
            return False
    return True


def rows_dataframe(
    section: dict,
    row_filter: dict | None = None,
    with_ids: bool = False,
    table_key: str | None = None,
):
    """Section rows -> display frame. With `table_key` the user's custom
    group/row ordering is applied (same ordering feeds the XLSX). When
    `with_ids` is set, returns (df, row_ids, effective_paths) - the paths
    drive the reorder controls."""
    hcols = hier_cols_for(section)
    kcols = key_cols_for(section)
    items = _section_order_items(section)
    ordered = _apply_custom_order(items, table_key) if table_key else [p for _, _, p in items]
    eff_of = {rid: path for path, rid, _ in items}
    recs, rids, paths = [], [], []
    for r in ordered:
        if hide_blk and r["type_raw"].split(".")[-1] == "BLK":
            continue
        if row_filter and not _row_in_filter(r, row_filter):
            continue
        vals = {c["inventory_id"]: r["values"].get(c["inventory_id"], "") for c in visible_cols}
        if focus_view and not any(v != "" for v in vals.values()):
            continue

        eff = eff_of.get(str(r["row_id"]), tuple(r["path"]))
        is_header = len(eff) == len(r["path"]) + 1  # a group/subgroup header row
        name = r["name"]
        if indent_names and r["path"]:
            name = ("\u00a0" * 4 * len(r["path"])) + name
        rec = {"Name": name}
        for i, hc in enumerate(hcols):
            v = r["path"][i] if len(r["path"]) > i else ""
            # a header row belongs to ITS OWN group/subgroup: show its name in
            # its level's cell so it merges (and moves) with its block
            if not v and is_header and i == len(r["path"]):
                v = r["name"]
            rec[hc] = v
        if show_type_col:
            rec["Row type"] = r["type"]
        for c, t in zip(visible_special_cols, special_tuples):
            rec[t] = r["values"].get(c["col_key"], "")
        for c, t in zip(visible_cols, col_tuples):
            rec[t] = vals[c["inventory_id"]]
        recs.append(rec)
        rids.append(str(r["row_id"]))
        paths.append(eff)
    # special (Display) columns sit BETWEEN the key block and the experiments
    df = pd.DataFrame(recs).reindex(columns=kcols + special_tuples + col_tuples).fillna("")
    return (df, rids, paths) if with_ids else df


def _merge_parents(names: list[str]) -> list[list[int]]:
    """Which already-merged columns each merge column depends on (its 'parents').

    Default hierarchy is strictly left-to-right, BUT interval columns are special:
    a time point (Interval 1/2) repeats across every property measured at it, so
    gating it behind Data Column / Unit would keep it from merging on a multi-
    property task. Instead an interval merges on its own repeated values, gated
    only by the OUTERMOST key (Data Template / Property Task) and any interval
    column to its left - so it still respects the top-level block boundary but
    spans across the different properties inside one block."""
    parents: list[list[int]] = []
    nonint: list[int] = []
    inte: list[int] = []
    for i, nm in enumerate(names):
        if str(nm).startswith("Interval "):
            parents.append(([nonint[0]] if nonint else []) + inte[:])
            inte.append(i)
        else:
            parents.append(list(nonint))
            nonint.append(i)
    return parents


def _merge_runs(
    rows: list[list[str]], n_merge: int, parents: list[list[int]] | None = None
) -> list[list[int]]:
    """For the first n_merge columns, compute the rowspan of each cell.
    span[r][c] = number of rows this cell spans (0 = absorbed by the cell above).
    `parents[c]` lists the columns that must still be merged for column c to keep
    a run going; when omitted it is every column to the left (a plain hierarchy),
    so 'Cobb Value' under a new Data Template starts a fresh merge."""
    n = len(rows)
    span = [[1] * n_merge for _ in range(n)]
    if parents is None:
        parents = [list(range(c)) for c in range(n_merge)]
    for c in range(n_merge):
        gate = parents[c]
        r = 0
        while r < n:
            k = r + 1
            while (
                k < n
                and rows[k][c] == rows[r][c]
                and rows[k][c] != ""
                and all(span[k][cc] == 0 for cc in gate)  # required parents still merged
            ):
                k += 1
            span[r][c] = k - r
            for j in range(r + 1, k):
                span[j][c] = 0
            r = k
    return span


def _merged_html(
    disp: pd.DataFrame,
    merge_cols: list[str],
    freeze: int,
    max_height: int = 620,
    key_width: int = 220,
    data_width: int = 150,
) -> str:
    """A real merged-cell table (HTML rowspan), like Excel's Merge & Center.
    Streamlit's grid cannot merge cells, so the merged view is rendered as HTML."""
    cols = list(disp.columns)
    n_merge = len([c for c in merge_cols if c in cols])
    merge_idx = [cols.index(c) for c in merge_cols if c in cols]
    body = disp.astype(str).values.tolist()
    # reorder so the merge columns come first in the span computation
    reord = merge_idx + [i for i in range(len(cols)) if i not in merge_idx]
    rows = [[r[i] for i in reord] for r in body]
    hdr = [cols[i] for i in reord]
    span = _merge_runs(rows, n_merge, _merge_parents(hdr[:n_merge]))

    css = (
        "<style>"
        # table-layout:fixed makes the colgroup widths authoritative, so the width
        # controls can BOTH grow and shrink a column (max-content only ever grew).
        ".mtbl{border-collapse:collapse;font-size:13px;table-layout:fixed}"
        ".mtbl th,.mtbl td{border:1px solid #d9d9d9;padding:5px 9px;vertical-align:middle;"
        "overflow:hidden;word-break:break-word}"
        ".mtbl th{background:#f2f2f2;position:sticky;top:0;z-index:3;text-align:left}"
        ".mtbl td.k{background:#fbfbfb;font-weight:500}"
        ".mtbl .stick{position:sticky;background:#fff;z-index:2}"
        ".mtbl th.stick{z-index:4;background:#f2f2f2}"
        ".mwrap{overflow:auto;max-height:" + str(max_height) + "px;"
        "border:1px solid #e6e6e6;border-radius:6px}"
        "</style>"
    )

    # sticky left offsets for the frozen columns; column widths are user-adjustable
    widths = [key_width if i < n_merge else data_width for i in range(len(hdr))]
    total_w = sum(widths)
    offs, acc = [], 0
    for w in widths:
        offs.append(acc)
        acc += w

    def cls(i):
        return " stick" if i < freeze else ""

    def sty(i):
        return f' style="left:{offs[i]}px;min-width:{widths[i]}px"' if i < freeze else ""

    h = [css, f'<div class="mwrap"><table class="mtbl" style="width:{total_w}px">']
    # <colgroup> sets an explicit width for every column so the width controls take
    # effect even for the non-frozen columns (which get no per-cell min-width).
    h.append("<colgroup>")
    for w in widths:
        h.append(f'<col style="width:{w}px;min-width:{w}px">')
    h.append("</colgroup>")
    h.append("<thead><tr>")
    for i, c in enumerate(hdr):
        h.append(f'<th class="{cls(i).strip()}"{sty(i)}>{c}</th>')
    h.append("</tr></thead><tbody>")
    for r in range(len(rows)):
        h.append("<tr>")
        for i, c in enumerate(hdr):
            if i < n_merge:
                s = span[r][i]
                if s == 0:
                    continue
                rs = f' rowspan="{s}"' if s > 1 else ""
                h.append(f'<td class="k{cls(i)}"{sty(i)}{rs}>{rows[r][i]}</td>')
            else:
                h.append(f'<td class="{cls(i).strip()}"{sty(i)}>{rows[r][i]}</td>')
        h.append("</tr>")
    h.append("</tbody></table></div>")
    return "".join(h)


def show_df(
    df: pd.DataFrame,
    key_labels: list[str],
    table_key: str,
    row_ids: list[str] | None = None,
    merge_cols: list[str] | None = None,
    order_paths: list[tuple] | None = None,
):
    """Render a section table with per-table controls.

    Column labels are SINGLE STRINGS - tuple (code, description) labels became a
    pandas MultiIndex, which Streamlit stringifies for long names, leaking the raw
    tuple repr into cells. The description rides along as a header tooltip and an
    optional first row.

    `order_paths` (one hierarchy path per row, aligned with `row_ids`) enables
    the reorder controls: an '↕ Reorder' popover to move whole group/subgroup
    blocks among their siblings, and an editable '#' column to move a single
    row INSIDE its own block (it can never leave the block).
    """
    if df.empty:
        st.info("No rows to display.")
        return

    # MERGE ORDER: outermost hierarchy first. 'Name' is the leaf (unique per row);
    # if it led the list, every run would break immediately and nothing would merge -
    # which is why Group/Subgroup never merged before.
    merge_cols = merge_cols or [k for k in key_labels if k != "Name"]

    row_ids = row_ids or [str(i) for i in range(len(df))]
    sel_key, applied_key = f"sel::{table_key}", f"applied::{table_key}"
    sel: dict[str, bool] = st.session_state.setdefault(sel_key, {})
    for rid in row_ids:
        sel.setdefault(rid, True)

    # ---- controls -----------------------------------------------------------
    # The three row-selection buttons are kept tight together on the left; the
    # view controls (freeze / merge / hide / full screen) follow after a gap.
    def _clear_editor_state() -> None:
        """Drop this table's data_editor widget state. The editor stores the
        user's clicks as POSITIONAL deltas against the input frame; when the
        selection is changed programmatically (Select all / Unselect all /
        Apply) those stale deltas would fight the new input - the checkbox
        appears to bounce back, which is exactly the 'hard to tick' symptom."""
        for _k in [k for k in st.session_state if str(k).startswith(f"ed::{table_key}")]:
            del st.session_state[_k]

    b_sel, c_ro, c_fz, c_mg, c_hide, c_full = st.columns(
        [2.0, 0.9, 1.2, 1.0, 1.9, 0.8]
    )
    with b_sel, _bordered():
        st.caption("Select visible rows")
        bb1, bb2, bb3 = st.columns(3)
        with bb1:
            if st.button("All", key=f"sa::{table_key}", use_container_width=True):
                for rid in row_ids:
                    sel[rid] = True
                st.session_state[applied_key] = False  # bring every row back into view
                _clear_editor_state()
                st.rerun()
        with bb2:
            if st.button("None", key=f"ua::{table_key}", use_container_width=True):
                for rid in row_ids:
                    sel[rid] = False
                _clear_editor_state()
                st.rerun()
        with bb3:
            if st.button(
                "Apply", key=f"ap::{table_key}", type="primary", use_container_width=True
            ):
                st.session_state[applied_key] = True
                _clear_editor_state()
                st.rerun()
    with c_ro:
        if order_paths is not None:
            with _popover("↕ Reorder"):
                st.caption(
                    "Move a whole group/block up or down among its own siblings - "
                    "every row of the block travels with it, and a block can never "
                    "leave its parent. To move a SINGLE row inside its block, edit "
                    "the **#** column directly in the table."
                )
                _nodes: list[tuple] = []
                _seen_nodes: set[tuple] = set()
                for _p in order_paths:
                    for _k in range(1, len(_p) + 1):
                        _n = _p[:_k]
                        if _n not in _seen_nodes:
                            _seen_nodes.add(_n)
                            _nodes.append(_n)
                if not _nodes:
                    st.caption("This table has no groups/blocks to reorder.")
                _sibs: dict[tuple, list[tuple]] = {}
                for _n in _nodes:
                    _sibs.setdefault(_n[:-1], []).append(_n)
                for _n in _nodes:
                    _sl = _sibs[_n[:-1]]
                    _i = _sl.index(_n)
                    _cc = st.columns([6, 1, 1])
                    _cc[0].markdown(
                        ("&nbsp;" * 6 * (len(_n) - 1)) + f"**{_n[-1]}**",
                        unsafe_allow_html=True,
                    )
                    _kb = hashlib.md5("¦".join(_n).encode()).hexdigest()[:10]
                    with _cc[1]:
                        if st.button("⬆", key=f"gu::{table_key}::{_kb}", disabled=_i == 0):
                            _swap_group_order(table_key, _n, _sl[_i - 1])
                            _clear_editor_state()
                            st.rerun()
                    with _cc[2]:
                        if st.button(
                            "⬇", key=f"gd::{table_key}::{_kb}", disabled=_i == len(_sl) - 1
                        ):
                            _swap_group_order(table_key, _n, _sl[_i + 1])
                            _clear_editor_state()
                            st.rerun()
    with c_fz:
        freeze = st.number_input(
            "Freeze columns",
            min_value=0,
            max_value=len(key_labels) + len(col_tuples) + len(special_tuples),
            value=len(key_labels),
            step=1,
            key=f"fz::{table_key}",
            help="How many columns stay pinned on the left while you scroll sideways.",
        )
    with c_mg:
        merge = st.checkbox(
            "Merge cells",
            value=False,
            key=f"mg::{table_key}",
            help="Combine repeated Group / Subgroup / Interval cells into one spanning "
            "cell (Excel-style). The merged view is read-only - untick to edit the "
            "row selection.",
        )
    # 'Hide columns' and 'Full screen' now apply in BOTH the interactive and the
    # merged views. The dropdown lists EVERY column - the key columns (Group /
    # Subgroup / Name / ...) as well as the experiment columns - so any of them can
    # be hidden.
    hideable = list(key_labels) + [t[0] for t in col_tuples + special_tuples]
    with c_hide:
        hidden_cols = st.multiselect(
            "Hide columns",
            hideable,
            key=f"hide::{table_key}",
            help="Hide any column from this table (key columns included).",
        )
    with c_full:
        full = st.checkbox(
            "Full screen",
            key=f"full::{table_key}",
            help="Expand this table to (almost) the full window height.",
        )
    hidden_set = set(hidden_cols)

    applied = st.session_state.get(applied_key, False)

    # ---- build the display frame --------------------------------------------
    disp = df.copy()
    disp.columns = [c if isinstance(c, str) else str(c[0]) for c in disp.columns]
    # Arrow needs one type per column (Trial is an int, the description row is text)
    disp = disp.astype(str).replace({"None": "", "nan": "", "<NA>": ""})
    disp.insert(0, "✓", [bool(sel.get(rid, True)) for rid in row_ids])
    disp["__rid__"] = row_ids

    if applied:
        disp = disp[disp["✓"]]
        if disp.empty:
            st.warning("No rows selected. Press **All** to bring them all back.")
            return

    # ---- '#' column: the row's position INSIDE its own block ------------------
    # Editable; typing a new number moves the row within its block only (the
    # editor callback permutes the block's ordinal values, so a row can never
    # land in another group/subgroup).
    if order_paths is not None:
        _path_of = dict(zip(row_ids, order_paths))
        _blocks: dict[tuple, list[str]] = {}
        for _rid in disp["__rid__"]:
            _blocks.setdefault(tuple(_path_of.get(_rid, ())), []).append(_rid)
        _pos_of = {
            _rid: _j + 1 for _blk in _blocks.values() for _j, _rid in enumerate(_blk)
        }
        disp.insert(1, "#", [float(_pos_of.get(_rid, 1)) for _rid in disp["__rid__"]])
        st.session_state[f"edblocks::{table_key}"] = {
            "blocks": {"¦".join(k): v for k, v in _blocks.items()},
            "of": {rid: "¦".join(k) for k, v in _blocks.items() for rid in v},
        }

    if show_desc_row and col_tuples:
        head = {c: "" for c in disp.columns}
        head["✓"] = False
        head[key_labels[0]] = "Description"
        head["__rid__"] = "__desc__"
        for code, desc in col_tuples + special_tuples:
            if code in disp.columns:
                head[code] = desc
        disp = pd.concat([pd.DataFrame([head]), disp], ignore_index=True)
        if "#" in disp.columns:
            disp["#"] = pd.to_numeric(disp["#"], errors="coerce")

    # ---- decimals: round the experiment (value) columns for display ----------
    if DECIMALS is not None:
        for _code, _ in col_tuples + special_tuples:
            if _code in disp.columns:
                disp[_code] = disp[_code].map(lambda x: _apply_decimals_text(x, DECIMALS))

    # ---- MERGED (read-only, real spanning cells) ------------------------------
    if merge:
        body = disp.drop(columns=[c for c in ("✓", "#", "__rid__") if c in disp.columns])
        shown = [c for c in body.columns if c not in hidden_set]
        body = body[shown]
        # Column widths are adjustable in the merged view.
        w1, w2 = st.columns(2)
        with w1:
            key_w = st.number_input(
                "Key column width (px)",
                min_value=60,
                max_value=600,
                value=220,
                step=10,
                key=f"kw::{table_key}",
                help="Width of the merged key columns (Group / Subgroup / Name / ...).",
            )
        with w2:
            dat_w = st.number_input(
                "Data column width (px)",
                min_value=50,
                max_value=400,
                value=150,
                step=10,
                key=f"dw::{table_key}",
                help="Width of the experiment (value) columns.",
            )
        st.markdown(
            _merged_html(
                body,
                [c for c in merge_cols if c in body.columns],
                int(freeze),
                max_height=880 if full else 620,
                key_width=int(key_w),
                data_width=int(dat_w),
            ),
            unsafe_allow_html=True,
        )
        st.caption(
            "Merged view is read-only. The XLSX export applies the same merges. "
            "Untick **Merge cells** to change the row selection."
        )
        return

    # ---- INTERACTIVE (checkboxes + pinned columns) ---------------------------
    ordered = [c for c in disp.columns if c != "__rid__" and c not in hidden_set]
    cfg: dict[str, Any] = {
        "✓": st.column_config.CheckboxColumn(
            "", help="Tick the rows to keep, then press Apply.", pinned=True
        ),
        "__rid__": None,  # hidden
    }
    if "#" in disp.columns:
        cfg["#"] = st.column_config.NumberColumn(
            "#",
            help="Position of the row INSIDE its own group/subgroup block. Type "
            "a new number to move the row within its block - it can never leave "
            "its block. Whole blocks are moved with the ↕ Reorder menu.",
            min_value=1,
            step=1,
            pinned=True,
        )
    for i, c in enumerate([x for x in ordered if x not in ("✓", "#")], start=1):
        desc = next((d for code, d in col_tuples + special_tuples if code == c), None)
        cfg[c] = st.column_config.Column(label=c, help=desc or None, pinned=i <= freeze)

    # RELIABLE TICK BOXES. Two flakiness sources are removed here:
    #   1. The editor stores clicks as POSITIONAL deltas - if the visible row
    #      set changes (filters, Apply) under a fixed widget key, old deltas
    #      land on the WRONG rows. The key therefore embeds a signature of the
    #      row order, so a changed row set always gets a fresh editor.
    #   2. Ticks were previously read back AFTER rendering, one run behind the
    #      click. An on_change callback now writes each click into the
    #      selection store BEFORE the rerun rebuilds the table, so a single
    #      click always sticks.
    _order = [str(x) for x in disp["__rid__"]]
    _sig = hashlib.md5(("\x1f".join(_order) + f"::{int(bool(applied))}").encode()).hexdigest()[:10]
    ed_key = f"ed::{table_key}::{_sig}"
    ord_key = f"edorder::{table_key}"
    st.session_state[ord_key] = _order
    # a changed signature orphans the previous editor state - prune it
    for _k in [
        k
        for k in st.session_state
        if str(k).startswith(f"ed::{table_key}::") and k != ed_key
    ]:
        del st.session_state[_k]

    def _sync_editor(_ed_key=ed_key, _sel_key=sel_key, _ord_key=ord_key, _tk=table_key) -> None:
        """Runs BEFORE the rerun: persists checkbox ticks and applies '#' row
        moves (permuting the ordinal VALUES of the row's own block only)."""
        state = st.session_state.get(_ed_key)
        order = st.session_state.get(_ord_key) or []
        if not isinstance(state, dict):
            return
        sel_map = st.session_state.setdefault(_sel_key, {})
        binfo = st.session_state.get(f"edblocks::{_tk}") or {}
        row_meta = st.session_state.get(f"ordmeta_rows::{_tk}") or {}
        rorder = st.session_state.setdefault(f"rorder::{_tk}", {})
        for pos, delta in (state.get("edited_rows") or {}).items():
            try:
                p = int(pos)
            except (TypeError, ValueError):
                continue
            if not (0 <= p < len(order)) or order[p] == "__desc__":
                continue
            rid = order[p]
            if "✓" in delta:
                sel_map[rid] = bool(delta["✓"])
            if "#" in delta and delta["#"] is not None:
                blk = (binfo.get("of") or {}).get(rid)
                rows_b = list((binfo.get("blocks") or {}).get(blk) or [])
                if rid not in rows_b:
                    continue
                try:
                    tgt = int(float(delta["#"]))
                except (TypeError, ValueError):
                    continue
                tgt = max(1, min(tgt, len(rows_b)))
                # current effective ordinals of the block, in display order
                vals = [float(rorder.get(r, row_meta.get(r, i))) for i, r in enumerate(rows_b)]
                rows_b.remove(rid)
                rows_b.insert(tgt - 1, rid)
                for v, r in zip(sorted(vals), rows_b):
                    rorder[r] = v

    # Only pass an explicit height for the full-screen view; omitting it lets the
    # grid auto-size (passing height=None raises on some Streamlit versions).
    editor_kwargs: dict[str, Any] = {"height": 800} if full else {}
    st.data_editor(
        disp,
        use_container_width=True,
        hide_index=True,
        column_config=cfg,
        column_order=ordered,
        disabled=[c for c in ordered if c not in ("✓", "#")],
        key=ed_key,
        on_change=_sync_editor,
        **editor_kwargs,
    )


# ===========================================================================
# 6) Results drill-down (lazy, per selected Property Task) + interval resolve
# ===========================================================================
def _column_value(col) -> str:
    """Recorded value of one PropertyValue. Albert stores it in the NESTED
    PropertyData object (col.property_data.value) - the top-level value/
    valueNumeric/valueString fields are usually empty on a GET."""
    pdat = getattr(col, "property_data", None)
    for src in (pdat, col):
        if src is None:
            continue
        for attr in ("value", "numeric_value", "string_value"):
            v = getattr(src, attr, None)
            if v not in (None, ""):
                return str(v)
    return ""


def _unit_name(col) -> str:
    u = getattr(col, "unit", None)
    if isinstance(u, dict):
        return str(u.get("name") or u.get("Name") or "")
    return str(getattr(u, "name", "") or "")


def _records_from_tpds(
    tpds,
    task_name: str = "",
    task_id: str = "",
    wf_of_block: dict[str, str] | None = None,
    task_workflows: list[str] | None = None,
) -> list[dict]:
    wf_of_block = wf_of_block or {}
    task_workflows = task_workflows or []
    recs: list[dict] = []
    for tpd in tpds:
        dt = getattr(tpd, "data_template", None)
        dt_name = getattr(dt, "name", None) or getattr(dt, "id", "") or "(no template)"
        dt_id = getattr(dt, "id", None) or ""
        # PropertyDataInventoryInformation exposes `.inventory_id` (alias "id")
        inv = getattr(tpd, "inventory", None)
        inv_id = getattr(inv, "inventory_id", None) or getattr(inv, "id", None)
        lot_id = getattr(inv, "lot_id", None) or ""
        block_id = getattr(tpd, "block_id", None) or ""
        # block -> workflow is authoritative; the links on the property data are a fallback
        wf = getattr(tpd, "initial_workflow", None) or getattr(tpd, "finial_workflow", None)
        wf_id = wf_of_block.get(block_id) or (getattr(wf, "id", None) or "")
        for interval in getattr(tpd, "data", None) or []:
            if getattr(interval, "void", False):
                continue
            raw_iv = getattr(interval, "interval_combination", "") or ""
            for trial in getattr(interval, "trials", None) or []:
                if getattr(trial, "void", False):
                    continue
                for col in getattr(trial, "data_columns", None) or []:
                    val = _column_value(col)
                    if val == "":
                        continue
                    _col_unit = getattr(col, "unit", None)
                    recs.append(
                        {
                            "task_id": getattr(tpd, "task_id", None) or task_id,
                            "block_id": block_id,
                            "task_name": task_name,
                            "workflow_id": wf_id,
                            "task_workflows": task_workflows,
                            # ids used by the DataTemplate-first filter; the
                            # unit SYMBOL is not here (property data carries
                            # only Unit.id) - it is joined from the DT
                            # definition at display time.
                            "dt_id": dt_id,
                            "dc_id": getattr(col, "id", None) or "",
                            "unit_id": (
                                _col_unit.get("id")
                                if isinstance(_col_unit, dict)
                                else getattr(_col_unit, "id", None)
                            )
                            or "",
                            "Data Template": dt_name,
                            "Data Column": getattr(col, "name", "") or "",
                            "Unit": _unit_name(col),
                            "Trial": getattr(trial, "visible_trial_number", None)
                            or getattr(trial, "trial_number", ""),
                            "raw_interval": raw_iv,
                            "inventory_id": inv_id,
                            "lot_id": lot_id,
                            "value": val,
                        }
                    )
    return recs


def _n_axes(records: list[dict]) -> int:
    """How many Interval columns this data actually needs. Never assume 2:
    a block you believe is single-axis can still emit crossed tokens."""
    n = 0
    for r in records:
        if "__error__" in r:
            continue
        n = max(n, len(ROW_TOKEN_RE.findall(str(r.get("raw_interval", "") or ""))))
    return max(1, n)


def resolve_intervals(records: list[dict]) -> None:
    """Attach 'Interval 1'..'Interval N' from the workflow's IntervalCombinations.
    Recomputed on every render, so a stale resolution can never survive."""
    cache: dict[str, dict[str, list[str]]] = st.session_state.setdefault("wf_intervals", {})
    unresolved: dict[str, str] = st.session_state.setdefault("wf_unresolved", {})

    wanted = sorted(
        {
            w
            for r in records
            for w in ([r.get("workflow_id", "")] + list(r.get("task_workflows") or []))
            if w and w not in cache
        }
    )
    if wanted:
        from concurrent.futures import ThreadPoolExecutor

        with ThreadPoolExecutor(max_workers=min(8, len(wanted))) as ex:
            for wid, m in zip(
                wanted, ex.map(lambda w: _workflow_interval_map(client, w), wanted)
            ):
                cache[wid] = m

    n_axes = max(_n_axes(records), N_AXES)
    for r in records:
        raw = str(r.get("raw_interval", "") or "")
        cands = [
            c for c in ([r.get("workflow_id", "")] + list(r.get("task_workflows") or [])) if c
        ]
        axes: list[str] = []
        if raw:
            for wid in cands:
                hit = cache.get(wid, {}).get(raw)
                if hit:
                    axes = list(hit)
                    break
            if not axes:
                unresolved[raw] = f"workflows tried: {cands or '(none on block or task)'}"
                axes = ROW_TOKEN_RE.findall(raw) or [raw]  # show the raw token, flagged
        for i in range(n_axes):
            r[f"Interval {i + 1}"] = axes[i] if i < len(axes) else ""
        r["_n_axes"] = n_axes


def load_target_results(_client: Albert, store: dict, tasks_to_fetch: list[dict]) -> None:
    """DataTemplate-first loader (flat parallel fan-out).

    IMPORTANT: `get_all_task_properties(task_id)` is NOT a single HTTP call - its
    SDK source is `check_for_task_data(task)` followed by a SEQUENTIAL loop of
    `get_task_block_properties(...)` over EVERY block/inventory combo of the task.
    So the old "one call per task" fetched all ~11 blocks of a task serially and
    then threw away the blocks we did not want.

    This version does the same work the SDK does, but (a) filtered to only the
    blocks that belong to the SELECTED Data Templates, and (b) fanned out FLAT:
    every surviving (task x block x inventory) combo across ALL target tasks is
    issued into ONE thread pool, so there is no per-task serial stall. The pool
    size is the 'Parallel requests' input. Errors are STORED per task.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    workers = int(st.session_state.get("fetch_workers", 16))

    # block -> workflow map per task (from the catalog's inline Blocks[]).
    wf_info = {
        t["id"]: (
            {b["block_id"]: b["workflow_id"] for b in t.get("blocks", []) if b["workflow_id"]},
            list(dict.fromkeys(b["workflow_id"] for b in t.get("blocks", []) if b["workflow_id"])),
        )
        for t in property_tasks
    }

    # Which blocks are actually wanted per task = union of the occurrences of the
    # SELECTED Data Templates. A target task may also hold OTHER blocks (other
    # DTs) we must not download. want=None (a task with no selected occurrence)
    # falls back to "no block restriction" so data is never silently dropped.
    _sels = st.session_state.get("dt_selectors") or []
    wanted_blocks: dict[str, set[str]] = {}
    for _s in _sels:
        for _o in dt_index.get(_s["dt_id"], {}).get("occurrences", []):
            wanted_blocks.setdefault(_o["task_id"], set()).add(_o["block_id"])

    # --- STEP 1: check_for_task_data per task (parallel) -> combo plan ---------
    # The bar renders in RESULTS_PROGRESS_SLOT - directly under the Load Data
    # button in section 2 - falling back to in-place rendering without it.
    _prog_host = RESULTS_PROGRESS_SLOT if RESULTS_PROGRESS_SLOT is not None else st
    prog = _prog_host.progress(0.0, text="Planning block/inventory combos...")

    def _check(task: dict):
        try:
            combos = _client.property_data.check_for_task_data(task_id=task["id"])
            return task, list(combos), None
        except Exception as e:  # noqa: BLE001
            return task, [], f"{type(e).__name__}: {e}"

    plan: list[tuple[dict, str, str, str | None]] = []
    check_errors: dict[str, str] = {}
    with ThreadPoolExecutor(max_workers=min(workers, max(1, len(tasks_to_fetch)))) as ex:
        for task, combos, err in ex.map(_check, tasks_to_fetch):
            if err:
                check_errors[task["id"]] = err
                continue
            want = wanted_blocks.get(task["id"])  # None -> no block restriction
            for c in combos:
                if not getattr(c, "data_exists", True):
                    continue
                bid = getattr(c, "block_id", "") or ""
                if want is not None and bid not in want:
                    continue  # block belongs to a non-selected DT - skip download
                plan.append(
                    (task, getattr(c, "inventory_id", None), bid, getattr(c, "lot_id", None))
                )

    # --- STEP 2: FLAT parallel fan-out of get_task_block_properties -----------
    per_task_tpds: dict[str, list] = {t["id"]: [] for t in tasks_to_fetch}
    per_task_err: dict[str, str] = {}
    total = len(plan)

    def _fetch(item: tuple[dict, str, str, str | None]):
        task, inv_id, bid, lot_id = item
        try:
            tpd = _client.property_data.get_task_block_properties(
                inventory_id=inv_id, task_id=task["id"], block_id=bid, lot_id=lot_id
            )
            return task["id"], tpd, None
        except Exception as e:  # noqa: BLE001
            return task["id"], None, f"{type(e).__name__}: {e}"

    done = 0
    if total:
        with ThreadPoolExecutor(max_workers=min(workers, total)) as ex:
            futures = [ex.submit(_fetch, it) for it in plan]
            for fut in as_completed(futures):
                tid, tpd, err = fut.result()
                if err:
                    per_task_err.setdefault(tid, err)
                elif tpd is not None:
                    per_task_tpds[tid].append(tpd)
                done += 1
                prog.progress(
                    done / total,
                    text=f"Fetching {done}/{total} block x inventory (x{workers} parallel)",
                )

    # --- STEP 3: flatten per task via the SAME _records_from_tpds, then store --
    for t in tasks_to_fetch:
        tid = t["id"]
        if tid in check_errors:
            store[tid] = [{"__error__": check_errors[tid], "task_id": tid}]
            continue
        wf_of_block, all_wf = wf_info.get(tid, ({}, []))
        recs = _records_from_tpds(
            per_task_tpds.get(tid, []),
            task_name=t["name"],
            task_id=tid,
            wf_of_block=wf_of_block,
            task_workflows=all_wf,
        )
        # a per-combo error with zero usable records is surfaced, not hidden
        if not recs and tid in per_task_err:
            store[tid] = [{"__error__": per_task_err[tid], "task_id": tid}]
        else:
            store[tid] = recs
    prog.empty()
    st.rerun()



def _loaded_records() -> list[dict]:
    return [
        r
        for recs in st.session_state.get(RESULTS_STORE_KEY, {}).values()
        for r in recs
        if "__error__" not in r
    ]


# Interval columns are generated from the DATA, not hard-coded to two. A block you
# believe is single-axis can still emit crossed tokens (ROW3XROW6), and Albert does
# not guarantee only two axes - so the column count follows the tokens.
N_AXES = max(2, _n_axes(_loaded_records()))
INTERVAL_KEYS = [f"Interval {i + 1}" for i in range(N_AXES)]
RESULT_KEYS = ["Data Template", "Data Column", "Unit"] + INTERVAL_KEYS + ["Trial"]
# Keys for the "Merge Results by DT" view: one row per Data Template / Data Column
# / Unit / Interval, pooled across every selected Property Block (Trial dropped, so
# repeated trials collapse into one cell via the aggregation choice).
MERGE_DT_KEYS = ["Data Template", "Data Column", "Unit"] + INTERVAL_KEYS


def _dt_selector_passes(r: dict, sels: list[dict]) -> bool:
    """DataTemplate-first display filter: a record survives when it satisfies
    at least ONE configured DT selector row -
      * its (task_id, block_id) is an occurrence of that row's DT (block ids
        repeat across tasks, so the PAIR is matched, never block_id alone);
      * its data column is among the row's selected Data Columns;
      * its resolved Interval 1 / Interval 2 is among the row's selected
        setpoints (iv1/iv2 are MULTI-selections; an empty list = '(any)').
    Requires resolve_intervals() to have run on the record."""
    for s in sels:
        if (r.get("task_id"), r.get("block_id")) not in s["occ"]:
            continue
        if r.get("dc_id") not in s["dc_ids"]:
            continue
        if s["iv1"] and r.get("Interval 1", "") not in s["iv1"]:
            continue
        if s["iv2"] and r.get("Interval 2", "") not in s["iv2"]:
            continue
        return True
    return False


def results_long_df(records: list[dict]) -> pd.DataFrame:
    """Tidy/long table - one row per recorded value (analysis-ready).
    The DT / Data Column / interval selection is applied HERE (rows), never to
    the experiment columns. Intervals are resolved BEFORE filtering because the
    interval choice compares against the resolved axis labels."""
    recs = [r for r in records if "__error__" not in r]
    if not recs:
        return pd.DataFrame()
    resolve_intervals(recs)
    sels = st.session_state.get("dt_selectors") or []
    recs = [r for r in recs if _dt_selector_passes(r, sels)]
    if not recs:
        return pd.DataFrame()
    df = pd.DataFrame(recs)
    # Unit SYMBOLS come from the DT definitions (property data has only ids)
    _sym_of: dict[str, str] = {}
    for s in sels:
        _sym_of.update({k: v for k, v in s["unit_symbol_of"].items() if v})
    _dc_col = df["dc_id"] if "dc_id" in df.columns else [""] * len(df)
    _u_col = df["Unit"] if "Unit" in df.columns else [""] * len(df)
    df["Unit"] = [_sym_of.get(dc) or u for dc, u in zip(_dc_col, _u_col)]
    df["Experiment"] = df["inventory_id"].map(
        lambda i: (invid_to_tuple.get(i) or ("", ""))[0] or _strip_inv(str(i or ""))
    )
    df["Experiment name"] = df["inventory_id"].map(
        lambda i: (invid_to_tuple.get(i) or ("", ""))[1]
    )
    df["Visible (passes filters)"] = df["inventory_id"].isin(invid_to_tuple)
    return df


def _agg_cell(values, mode: str) -> str:
    """Combine several measurements that land in one (property x experiment) cell.

    mode 'avg' -> numeric mean of the measurements (e.g. 6.12, 6.65, 5.71 -> 6.16),
                  formatted to the same number of decimals as the inputs; falls back
                  to listing when the values are not all numeric.
    otherwise  -> the distinct values joined with ' | ' (original behaviour)."""
    vals = [str(x) for x in values if str(x) != ""]
    if not vals:
        return ""
    if mode == "avg":
        nums = []
        for x in vals:
            try:
                nums.append(float(x.replace(",", ".")))
            except ValueError:
                nums = []
                break  # non-numeric column -> list instead of averaging
        if nums:
            m = sum(nums) / len(nums)
            # keep as many decimals as the inputs carry (handles 6,12 and 6.12)
            decs = [len(x.replace(",", ".").split(".", 1)[1]) for x in vals if ("." in x or "," in x)]
            dec = min(max(decs, default=0), 6)
            return f"{m:.{dec}f}"
    return " | ".join(dict.fromkeys(vals))


def results_drilldown_df(
    records: list[dict],
    include_foreign: bool = False,
    group_keys: list[str] | None = None,
    keep_all_rows: bool = False,
    order_table_key: str | None = None,
) -> pd.DataFrame:
    """Pivot: DT | DC | Unit | I1 | I2 | Trial rows x visible experiment cols.
    `include_foreign` also shows inventory items filtered out or belonging to
    other sheets (dropped silently before = looked like 'no data').
    `keep_all_rows` keeps a property row even when none of the visible (passing)
    formulations carry a value for it, so the table stays visible under an active
    Advanced filter instead of collapsing to an empty frame.
    `group_keys` overrides the row key (e.g. MERGE_DT_KEYS drops Trial so several
    Property Blocks pool into one row per Data Template / Column / Interval)."""
    keys = group_keys or RESULT_KEYS
    long = results_long_df(records)
    if long.empty:
        return pd.DataFrame()
    agg_mode = st.session_state.get("results_agg_mode", "list")

    tuple_of = dict(invid_to_tuple)
    extra_cols: list[tuple[str, str]] = []
    if include_foreign:
        for inv in long.loc[~long["Visible (passes filters)"], "inventory_id"].dropna().unique():
            t = (_strip_inv(str(inv)), "(filtered out / other sheet)")
            if t not in extra_cols:
                extra_cols.append(t)
            tuple_of[inv] = t

    g = (
        long.groupby(keys + ["inventory_id"], dropna=False, sort=False)["value"]
        .apply(lambda v: _agg_cell(v, agg_mode))
        .reset_index()
    )
    recs = []
    for kv, chunk in g.groupby(keys, dropna=False, sort=False):
        rec = dict(zip(keys, kv if isinstance(kv, tuple) else (kv,)))
        has = False
        for _, r in chunk.iterrows():
            t = tuple_of.get(r["inventory_id"])
            if t:
                rec[t] = r["value"]
                has = True
        if has or keep_all_rows:
            recs.append(rec)
    if not recs:
        return pd.DataFrame()
    if order_table_key:
        # DataTemplate/DataColumn block ordering (same mechanism as the section
        # tables): rows are grouped DT -> Data Column and the user's custom
        # block/row order is applied on top. A row can never leave its DT/DC.
        items = [
            (
                (str(rec.get("Data Template", "")), str(rec.get("Data Column", ""))),
                "|".join(str(rec.get(k, "")) for k in keys),
                rec,
            )
            for rec in recs
        ]
        recs = _apply_custom_order(items, order_table_key)
    else:
        # Keep rows of the same Data Template together (the source order can
        # interleave them, e.g. a Coating Weight row between two Cobb Value
        # rows), so the merged Data Template cell spans them. First-appearance
        # order and the order within a template are both preserved (stable sort).
        dt_order: dict[str, int] = {}
        for rec in recs:
            dt_order.setdefault(str(rec.get("Data Template", "")), len(dt_order))
        recs.sort(key=lambda rec: dt_order[str(rec.get("Data Template", ""))])
    return pd.DataFrame(recs).reindex(columns=keys + col_tuples + extra_cols).fillna("")


# ===========================================================================
# 7) Render sections (all obey the same visible_cols)
# ===========================================================================
for s in sections:
    st.subheader(s["label"])

    if s["attr"] != "result_design":
        # --- hierarchy provenance: never let an inferred tree pass as fact ---
        if s["max_depth"] == 0:
            st.warning(
                "⚠️ **Group / Subgroup hierarchy unavailable for this section.** "
                f"`GET /api/v3/worksheet/design/.../rows/sequence` returned no tree"
                + (f" ({s['hierarchy_error']})" if s["hierarchy_error"] else "")
                + ". The flat grid response carries no parent, child, depth or "
                "indent field, so depth cannot be recovered from it - a BLK row "
                "followed by another BLK row is ambiguous between *child* and "
                "*sibling*. Rather than guess, the Group columns are left out. "
                "See the Diagnostics panel for the raw payload."
            )
        elif "ONE LEVEL" in s["hierarchy_source"]:
            st.info(
                "ℹ️ Only a single Group level is available for this section "
                "(the sequence endpoint returned no nested subgroups)."
            )

        # --- per-level row filters (Group, Subgroup 1, ...) ------------------
        row_filter: dict[int, list[str]] = {}
        hcols = hier_cols_for(s)
        if hcols:
            fcols = st.columns(len(hcols))
            for lv, (hc, fc) in enumerate(zip(hcols, fcols)):
                opts = sorted(
                    {r["path"][lv] for r in s["rows"] if len(r["path"]) > lv and r["path"][lv]}
                )
                # rows with nothing at this level (top-level headers, shallow
                # branches) - selectable via (None) instead of vanishing
                if any(len(r["path"]) <= lv for r in s["rows"]):
                    opts = [NONE_LABEL] + opts
                with fc:
                    row_filter[lv] = st.multiselect(
                        hc,
                        opts,
                        key=f"rowfilter::{s['attr']}::{lv}",
                        help=f"'{NONE_LABEL}' = rows with no {hc.lower()}.",
                    )

        sdf, srids, spaths = rows_dataframe(
            s, row_filter, with_ids=True, table_key=f"sec::{s['attr']}"
        )
        show_df(
            sdf,
            key_cols_for(s),
            table_key=f"sec::{s['attr']}",
            row_ids=srids,
            order_paths=spaths,
        )
        continue

    # ----- Results: DataTemplate-first. What gets downloaded is defined by the
    # "Selection of Data Templates in Results" panel in section 2, and ALL the
    # load controls (checkboxes, aggregation, Load Data / Reload / parallel
    # requests) also live there now - here only the fetch is executed and ONE
    # merged table is rendered.
    _sels = dt_selectors
    store = results_store
    if _load_clicked and _results_to_fetch:
        load_target_results(client, store, _results_to_fetch)
        # (ends in st.rerun(); nothing to fetch -> fall through and render)

    # --- render: ONE table, always merged by DT -------------------------------
    loaded_recs = []
    _errors: list[str] = []
    for _tid in target_tasks:
        for _r in store.get(_tid, []):
            if "__error__" in _r:
                _errors.append(f"{_tid}: {_r['__error__']}")
            else:
                loaded_recs.append(_r)
    if _errors:
        st.error(f"API call failed: {_errors[0]}")

    _loaded_n = sum(1 for _tid in target_tasks if _tid in store)
    if _sels and not _loaded_n:
        st.caption("No property data loaded yet - press **Load Data** (section 2).")
    elif _sels:
        st.caption(
            f"Merged view · {_loaded_n} task(s) loaded · one row per "
            "Data Template / Data Column / Interval, filtered by the DT panel "
            "in section 2 (Data Columns + intervals are display filters on the "
            "cached download)."
        )

    if long_view:
        ldf = results_long_df(loaded_recs)
        if ldf.empty:
            st.info("No rows match the current DT / Data Column / interval selection.")
        else:
            st.dataframe(ldf, use_container_width=True, hide_index=True)
    else:
        mdf = results_drilldown_df(
            loaded_recs, include_foreign=include_foreign, group_keys=MERGE_DT_KEYS,
            order_table_key="res::merged_by_dt",
        )
        if mdf.empty and len(visible_cols) < len(exp_cols_all):
            # keep the table visible when column filters hid every carrier
            mdf = results_drilldown_df(
                loaded_recs, include_foreign=include_foreign,
                group_keys=MERGE_DT_KEYS, keep_all_rows=True,
                order_table_key="res::merged_by_dt",
            )
        if mdf.empty:
            if loaded_recs:
                st.info(
                    "Data is loaded, but no row matches the current Data Column / "
                    "interval selection - or none of it belongs to the visible "
                    "experiments (tick 'Include experiments filtered out')."
                )
        else:
            rids = [
                "|".join(str(mdf.iloc[i][k]) for k in MERGE_DT_KEYS) for i in range(len(mdf))
            ]
            # reorder blocks = Data Template › Data Column (rows can never
            # leave their DT / Data Column)
            rpaths = [
                (str(mdf.iloc[i]["Data Template"]), str(mdf.iloc[i]["Data Column"]))
                for i in range(len(mdf))
            ]
            show_df(
                mdf,
                MERGE_DT_KEYS,
                table_key="res::merged_by_dt",
                row_ids=rids,
                order_paths=rpaths,
            )
    # (the DT index & Load plan diagnostics expander moved to section 2, next
    #  to the load controls)


# ===========================================================================
# 8) Downloads (respect the global filters)
# ===========================================================================
st.header("4️⃣ Download")


def all_results_df() -> pd.DataFrame:
    store = st.session_state.get(RESULTS_STORE_KEY, {})
    frames = []
    for task_id, recs in store.items():
        clean = [r for r in recs if "__error__" not in r]
        df = results_drilldown_df(clean, include_foreign=True)
        if not df.empty:
            df.insert(0, "Property Task", clean[0].get("task_name") or task_id)
            frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def all_results_long_df() -> pd.DataFrame:
    store = st.session_state.get(RESULTS_STORE_KEY, {})
    frames = [results_long_df(recs) for recs in store.values()]
    frames = [f for f in frames if not f.empty]
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True)
    cols = [
        "task_id", "task_name", "block_id", "workflow_id",
        "Data Template", "Data Column", "Unit",
        *INTERVAL_KEYS, "raw_interval", "Trial",
        "Experiment", "Experiment name", "inventory_id", "lot_id",
        "Visible (passes filters)", "value",
    ]
    return out.reindex(columns=[c for c in cols if c in out.columns])


def build_xlsx() -> bytes:
    """Report-ready workbook.

    THE FIX: every section previously wrote its data columns at its own offset
    (Product started after its key columns, Results after 'Property Task' + 6 more),
    so the experiment columns did not line up down the page. Now there is ONE fixed
    grid: a key block of KEY_W columns on the left, then the experiment columns at
    the SAME absolute position for every section. Read straight down column F and
    you are reading one experiment across Product, Process, Results and Apps.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # --- mirror the ON-SCREEN state of every table ----------------------------
    # The workbook reproduces the Streamlit tables 1:1 - visible rows (the
    # per-level row filters + the ticked/applied row selection), visible
    # columns (each table's 'Hide columns' + the 'Display' menu + the global
    # filters), MERGED vs NON-MERGED cells (each table's own 'Merge cells'
    # checkbox) and the Decimals rounding.
    def _table_key_of(attr: str) -> str:
        return "res::merged_by_dt" if attr == "result_design" else f"sec::{attr}"

    hidden_of = {
        s["attr"]: set(st.session_state.get(f"hide::{_table_key_of(s['attr'])}") or [])
        for s in sections
    }
    merge_flag_of = {
        s["attr"]: bool(st.session_state.get(f"mg::{_table_key_of(s['attr'])}", False))
        for s in sections
    }

    # --- one key block wide enough for the widest section ---------------------
    # DataTemplate-first: the Results view is ALWAYS merged by DT. Key columns
    # hidden on screen are dropped from that section's key block.
    per_section_keys = {s["attr"]: key_cols_for(s) for s in sections}
    per_section_keys["result_design"] = list(MERGE_DT_KEYS)
    per_section_keys = {
        attr: [k for k in keys if k not in hidden_of.get(attr, set())]
        for attr, keys in per_section_keys.items()
    }
    KEY_W = max(1, max(len(v) for v in per_section_keys.values()))
    FIRST_EXP = KEY_W + 1  # 1-based column of the first experiment

    # Value columns exported = every special/experiment column visible in AT
    # LEAST one section table (special Display columns first, mirroring the
    # on-screen order). A column hidden in one specific table gets empty cells
    # in that section only - the workbook keeps ONE aligned column layout.
    _all_tuples = special_tuples + col_tuples
    EXPORT_TUPLES = [
        t for t in _all_tuples if any(t[0] not in hidden_of[s["attr"]] for s in sections)
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"

    NAVY = "1F3864"
    GREY = "F2F2F2"
    BAND = "DDEBF7"
    bold_w = Font(bold=True, color="FFFFFF", size=11)
    bold = Font(bold=True)
    ital = Font(italic=True, size=9, color="555555")
    sect_fill = PatternFill("solid", fgColor=NAVY)
    head_fill = PatternFill("solid", fgColor=GREY)
    band_fill = PatternFill("solid", fgColor=BAND)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- title block ----------------------------------------------------------
    ws.cell(row=1, column=1, value=f"Albert Worksheet - {TITLE_PROJECTS}").font = Font(
        bold=True, size=14
    )
    ws.cell(
        row=2,
        column=1,
        value=f"Sheet(s): {TITLE_SHEETS}   |   {len(visible_cols)} of {len(exp_cols_all)} "
        f"experiments shown   |   exported {pd.Timestamp.now():%Y-%m-%d %H:%M}",
    ).font = ital

    # --- frozen experiment header (ID over description) -----------------------
    HDR = 4
    for j, (code, desc) in enumerate(EXPORT_TUPLES):
        c1 = ws.cell(row=HDR, column=FIRST_EXP + j, value=code)
        c1.font, c1.alignment, c1.fill, c1.border = bold, ctr, band_fill, box
        if show_desc_row:  # description row mirrors the on-screen toggle
            c2 = ws.cell(row=HDR + 1, column=FIRST_EXP + j, value=desc)
            c2.font, c2.alignment, c2.border = ital, ctr, box
    ws.cell(row=HDR, column=1, value="Experiment →").font = bold
    ws.freeze_panes = ws.cell(row=HDR + 2, column=FIRST_EXP)

    r = HDR + 2

    def write_section(
        label: str,
        keys: list[str],
        rows_iter,
        merge_cols: list[str],
        extra_tuples: list[tuple[str, str]] | None = None,
    ) -> None:
        """One section table. `rows_iter` yields (keyvals, expvals) where
        expvals is already aligned to EXPORT_TUPLES (+ extra_tuples appended,
        e.g. Results' foreign-experiment columns)."""
        nonlocal r
        extra_tuples = extra_tuples or []
        n_vals = len(EXPORT_TUPLES) + len(extra_tuples)
        r += 1
        # full-width section banner
        ws.cell(row=r, column=1, value=label.upper()).font = bold_w
        for cc in range(1, FIRST_EXP + n_vals):
            ws.cell(row=r, column=cc).fill = sect_fill
        r += 1
        for i, k in enumerate(keys):
            c = ws.cell(row=r, column=1 + i, value=k)
            c.font, c.fill, c.border, c.alignment = bold, head_fill, box, lft
        for j in range(n_vals):
            c = ws.cell(row=r, column=FIRST_EXP + j)
            c.fill, c.border = head_fill, box
        # extra columns exist only in this section - header them here
        for j, (code, _desc) in enumerate(extra_tuples):
            c = ws.cell(row=r, column=FIRST_EXP + len(EXPORT_TUPLES) + j, value=code)
            c.font, c.alignment = bold, ctr
        r += 1

        first_data = r
        keymat: list[list[str]] = []
        for keyvals, expvals in rows_iter:
            padded = [str(keyvals[i]) if i < len(keyvals) else "" for i in range(KEY_W)]
            keymat.append(padded)
            for i in range(KEY_W):
                c = ws.cell(row=r, column=1 + i, value=padded[i])
                c.border, c.alignment = box, lft
            for j, v in enumerate(expvals):
                c = ws.cell(row=r, column=FIRST_EXP + j, value=_num(v))
                c.border, c.alignment = box, ctr
            r += 1

        # --- real Excel merges on the key columns (same runs as the UI) --------
        idxs = [keys.index(m) for m in merge_cols if m in keys]
        if keymat and idxs:
            ordered = [[row[i] for i in idxs] for row in keymat]
            spans = _merge_runs(ordered, len(idxs), _merge_parents([keys[i] for i in idxs]))
            for rr in range(len(keymat)):
                for cc, col_i in enumerate(idxs):
                    s = spans[rr][cc]
                    if s > 1:
                        ws.merge_cells(
                            start_row=first_data + rr,
                            start_column=1 + col_i,
                            end_row=first_data + rr + s - 1,
                            end_column=1 + col_i,
                        )
                        mc = ws.cell(row=first_data + rr, column=1 + col_i)
                        mc.alignment = Alignment(
                            horizontal="left", vertical="center", wrap_text=True
                        )

    def _num(v):
        """Write numbers as numbers so Excel can chart/aggregate them."""
        if isinstance(v, str):
            t = v.strip().replace(",", ".")
            try:
                return float(t) if t not in ("", "-") else v
            except ValueError:
                return v
        return v

    def _apply_row_selection(df: pd.DataFrame, rids: list[str], table_key: str) -> pd.DataFrame:
        """Export exactly what's on screen: if the user pressed Apply selection,
        only the ticked rows go into the workbook."""
        if not st.session_state.get(f"applied::{table_key}", False):
            return df
        sel = st.session_state.get(f"sel::{table_key}", {})
        keep = [i for i, rid in enumerate(rids) if sel.get(rid, True)]
        return df.iloc[keep]

    def _exp_values(row, hidden: set[str]) -> list:
        """One row's exported value cells: aligned to EXPORT_TUPLES, blanked
        where this section's table hides the column, rounded like the screen."""
        return [
            "" if t[0] in hidden else _apply_decimals_text(row.get(t, ""), DECIMALS)
            for t in EXPORT_TUPLES
        ]

    for s in sections:
        attr = s["attr"]
        hidden = hidden_of[attr]
        merge_this = merge_flag_of[attr]
        if attr != "result_design":
            keys = per_section_keys[attr]
            # same per-level Group/Subgroup row filters as the on-screen table
            row_filter = {
                lv: st.session_state.get(f"rowfilter::{attr}::{lv}") or []
                for lv in range(len(hier_cols_for(s)))
            }
            df, rids, _spaths = rows_dataframe(
                s, row_filter, with_ids=True, table_key=f"sec::{attr}"
            )
            df = _apply_row_selection(df, rids, f"sec::{attr}")
            write_section(
                s["label"],
                keys,
                (
                    ([row[k] for k in keys], _exp_values(row, hidden))
                    for _, row in df.iterrows()
                ),
                # merged cells only when the table's 'Merge cells' is ticked
                merge_cols=[k for k in keys if k != "Name"] if merge_this else [],
            )
        else:
            # EXACTLY the frame shown on screen: same records (the DT panel's
            # target tasks), same include_foreign choice, same fallback that
            # keeps property rows visible when column filters hid every carrier.
            rdf = results_drilldown_df(
                loaded_recs, include_foreign=include_foreign, group_keys=MERGE_DT_KEYS,
                order_table_key="res::merged_by_dt",
            )
            if rdf.empty and len(visible_cols) < len(exp_cols_all):
                rdf = results_drilldown_df(
                    loaded_recs,
                    include_foreign=include_foreign,
                    group_keys=MERGE_DT_KEYS,
                    keep_all_rows=True,
                    order_table_key="res::merged_by_dt",
                )
            extra_tuples: list[tuple[str, str]] = []
            if not rdf.empty:
                # the on-screen ticked/applied row selection
                rrids = [
                    "|".join(str(rdf.iloc[i][k]) for k in MERGE_DT_KEYS)
                    for i in range(len(rdf))
                ]
                rdf = _apply_row_selection(rdf, rrids, "res::merged_by_dt")
                # foreign-experiment columns (only present with include_foreign)
                extra_tuples = [
                    t
                    for t in rdf.columns
                    if isinstance(t, tuple) and t not in _all_tuples
                ]
            keys = per_section_keys["result_design"]
            write_section(
                s["label"],
                keys,
                (
                    (
                        [row.get(k, "") for k in keys],
                        _exp_values(row, hidden)
                        + [
                            _apply_decimals_text(row.get(t, ""), DECIMALS)
                            for t in extra_tuples
                        ],
                    )
                    for _, row in rdf.iterrows()
                )
                if not rdf.empty
                else iter(()),
                # Merge on all key columns - but only when the on-screen table
                # has 'Merge cells' ticked.
                merge_cols=keys if merge_this else [],
                extra_tuples=extra_tuples,
            )

    # --- widths ---------------------------------------------------------------
    ws.column_dimensions["A"].width = 34
    for i in range(2, KEY_W + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    for j in range(len(EXPORT_TUPLES)):
        ws.column_dimensions[get_column_letter(FIRST_EXP + j)].width = 16
    if show_desc_row:
        ws.row_dimensions[HDR + 1].height = 42

    # --- tidy long results on a second sheet (analysis-ready) -----------------
    ldf = all_results_long_df()
    if not ldf.empty:
        ws2 = wb.create_sheet("Results (long)")
        ws2.append(list(ldf.columns))
        for c in ws2[1]:
            c.font, c.fill, c.border = bold, head_fill, box
        for _, row in ldf.iterrows():
            ws2.append([_num(v) for v in row.tolist()])
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions
        for i, cname in enumerate(ldf.columns, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = max(
                12, min(38, len(str(cname)) + 4)
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


d1, d2, d3 = st.columns(3)
with d1:
    _xlsx = build_xlsx()
    if os.environ.get("ALBERT_DUMP_XLSX"):  # offline inspection / tests
        with open(os.environ["ALBERT_DUMP_XLSX"], "wb") as _f:
            _f.write(_xlsx)
    st.download_button(
        "📥 XLSX (filtered worksheet + results)",
        data=_xlsx,
        file_name=f"albert_{COMPARE_TAG}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with d2:
    rdf = all_results_df()
    st.download_button(
        "📥 CSV (results pivot)",
        data=rdf.to_csv(index=False) if not rdf.empty else "",
        file_name=f"albert_{COMPARE_TAG}_results.csv",
        mime="text/csv",
        use_container_width=True,
        disabled=rdf.empty,
    )
with d3:
    ldf = all_results_long_df()
    st.download_button(
        "📥 CSV (results tidy/long)",
        data=ldf.to_csv(index=False) if not ldf.empty else "",
        file_name=f"albert_{COMPARE_TAG}_results_long.csv",
        mime="text/csv",
        use_container_width=True,
        disabled=ldf.empty,
        help="One row per value, with resolved intervals - ready for pandas/PSD analysis.",
    )


# ===========================================================================
# Diagnostics
# ===========================================================================
with st.expander("🔧 Row hierarchy - raw `rows/sequence` payload & resolved paths"):
    for s in sections:
        st.markdown(f"**{s['label']}**")
        st.write(
            {
                "source": s["hierarchy_source"],
                "max depth (ancestor levels)": s["max_depth"],
                "JSON keys actually found": s["hierarchy_keys"] or "(none)",
                "ancestor rowIds with no name found": s.get("hierarchy_unresolved") or "(none)",
                "error": s["hierarchy_error"] or "(none)",
            }
        )
        if s["hierarchy_raw"] is not None:
            st.json(s["hierarchy_raw"], expanded=False)
        paths_df = pd.DataFrame(
            [
                {
                    "row_id": r["row_id"],
                    "type": r["type_raw"].split(".")[-1],
                    "name": r["name"],
                    "depth": r["depth"],
                    "breadcrumb": " > ".join(r["path"] + [r["name"]]),
                }
                for r in s["rows"]
            ]
        )
        st.dataframe(paths_df, use_container_width=True, hide_index=True)
        st.divider()

with st.expander("🔧 Interval resolution (token → setpoint)"):
    unres = st.session_state.get("wf_unresolved", {})
    if unres:
        st.error(
            "**Tokens that could not be resolved** (shown raw in the tables). "
            "Each entry lists the workflows that were searched:"
        )
        st.write(unres)
    else:
        st.success("Every interval token resolved to a setpoint.")
    st.caption(
        f"Interval columns in use: {N_AXES}. Tokens come from "
        "`Workflow.IntervalCombinations[].interval`; the setpoints and their "
        "left-to-right order come from the matching `intervalString`."
    )
    maps = st.session_state.get("wf_intervals", {})
    if maps:
        st.dataframe(
            pd.DataFrame(
                [
                    {"workflow": w, "token": tok, **{f"Interval {i+1}": a for i, a in enumerate(axes)}}
                    for w, m in maps.items()
                    for tok, axes in m.items()
                ]
            ),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.write("(no workflows loaded yet)")

with st.expander("🔧 Filter sources (facets, tags, predecessor, data templates)"):
    st.write("**Albert inventory facets** (`inventory.get_all_facets`, project-scoped):")
    st.dataframe(
        pd.DataFrame(
            [
                {"facet parameter": p, "value": n, "count": c}
                for p, vals in facets.items()
                for n, c in vals
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )
    st.write(
        "**Per-formulation filter data.** Tags come from `InventoryItem.tags` - the "
        "name is on `Tag.tag`, not `Tag.name` - with any id-only tag resolved via "
        "`tags.get_by_ids`. Predecessor is read from the Apps **PDC** row, the only "
        "place Albert stores it (it is not a field on InventoryItem)."
    )
    st.dataframe(
        pd.DataFrame(
            [
                {
                    "inventory_id": c["inventory_id"],
                    "column_id (Product Design)": c["column_id"],
                    "tags": ", ".join(inv_meta.get(c["inventory_id"], {}).get("tags", [])),
                    "predecessor": inv_meta.get(c["inventory_id"], {}).get("predecessor", ""),
                    "created_by": inv_meta.get(c["inventory_id"], {}).get("created_by", ""),
                    "data_templates": ", ".join(
                        sorted(dts_of_inv.get(c["inventory_id"], set()))
                    ),
                    "locked": c["locked"],
                }
                for c in exp_cols_all
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )
    st.write("**Data template id → name used in the filter:**", dt_name_of or "(none)")

with st.expander("🔧 Raw property-data payload (one task)"):
    if property_tasks:
        pick = st.selectbox(
            "Task", [f"{t['name']}  [{t['id']}]" for t in property_tasks], key="dbg_task"
        )
        tid = pick.split("[")[-1].rstrip("]")
        if st.button("Show raw response"):
            try:
                checks = client.property_data.check_for_task_data(task_id=tid)
                st.write("**check_for_task_data** (drives which combos are fetched):")
                st.dataframe(
                    pd.DataFrame([c.model_dump() for c in checks]), use_container_width=True
                )
                tpds = client.property_data.get_all_task_properties(
                    task_id=tid, with_data_only=True
                )
                st.write(f"**get_all_task_properties** -> {len(tpds)} block/inventory objects")
                for tpd in tpds[:3]:
                    st.json(tpd.model_dump(by_alias=True, mode="json"), expanded=False)
            except Exception as e:  # noqa: BLE001
                st.exception(e)

with st.expander("🔧 Diagnostics"):
    st.write("**Columns** (is_label_col=True are excluded as duplicates of row names):")
    st.dataframe(pd.DataFrame(columns), use_container_width=True, hide_index=True)
    st.write("**Formulation metadata used by the filters** (tags / predecessor / creator):")
    st.dataframe(
        pd.DataFrame(
            [
                {
                    "inventory_id": k,
                    "tags": ", ".join(v["tags"]),
                    "predecessor": v["predecessor"],
                    "created_by": v["created_by"],
                    "name": v["name"],
                    "alias": v["alias"],
                }
                for k, v in inv_meta.items()
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )
    st.write("**Property Tasks & their Data Templates** (drives the DT filter):")
    st.dataframe(
        pd.DataFrame(
            [
                {
                    "task": t["name"],
                    "id": t["id"],
                    "state": t["state"],
                    "data_templates": ", ".join(t["data_templates"]),
                    "n_inventories": len(t["inventory_ids"]),
                }
                for t in property_tasks
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )
    for s in sections:
        st.write(
            f"**{s['label']}** row types:",
            sorted({r["type_raw"] for r in s["rows"]}),
            " | link_id samples:",
            [r["link_id"] for r in s["rows"][:8]],
        )


# ---------------------------------------------------------------------------
# TEMPORARY PERF PROBE  (delete once the pydantic-vs-json question is settled)
# Measures the cost of the SDK's typed model (pydantic) against a raw HTTP GET
# parsed with plain JSON, on the same task, over the same authenticated session.
# ---------------------------------------------------------------------------
with st.expander("⏱️ Perf probe (temporary)"):
    import time as _time
    import statistics as _stat
    import inspect as _inspect

    _tid = st.text_input("Task id to probe", value="TASFOR969623", key="perf_tid")

    if st.button("Measure pydantic vs json", key="perf_go"):
        # Show the real method source -> reveals the exact REST path the SDK uses.
        try:
            st.code(
                _inspect.getsource(client.property_data.get_all_task_properties),
                language="python",
            )
        except Exception as _e:  # noqa: BLE001
            st.write("(could not read method source:", _e, ")")

        # Same authenticated session the app already uses for raw workflow reads.
        _raw_params = {"taskId": _tid, "withDataOnly": "false"}

        def _raw_get():
            r = client.session.get("/api/v3/propertydata", params=_raw_params)
            r.raise_for_status()
            return r.json()  # plain dict, no pydantic

        def _best_of(fn, n=5):
            ts = []
            for _ in range(n):
                _t0 = _time.perf_counter()
                fn()
                ts.append(_time.perf_counter() - _t0)
            return min(ts), _stat.median(ts)

        try:
            # Warm-up (open connection / lazy auth) - not timed.
            client.property_data.get_all_task_properties(task_id=_tid)
            _raw_get()

            _sdk_min, _sdk_med = _best_of(
                lambda: client.property_data.get_all_task_properties(task_id=_tid)
            )
            _raw_min, _raw_med = _best_of(_raw_get)

            st.success(
                f"SDK + pydantic : min {_sdk_min:.3f}s   median {_sdk_med:.3f}s\n"
                f"raw + json     : min {_raw_min:.3f}s   median {_raw_med:.3f}s\n"
                f"RATIO (median) : {_sdk_med / _raw_med:.1f}x"
            )
        except Exception as _e:  # noqa: BLE001
            st.error(
                "Raw GET failed - most likely the inferred path or param name is "
                f"wrong.\n{type(_e).__name__}: {_e}"
            )
            st.info(
                "Copy the method source shown above and send it back - the exact "
                "path is in there."
            )
